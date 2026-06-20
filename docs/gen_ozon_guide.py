"""生成 OZON 模块详细说明文档"""
import openpyxl, os
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
hf = Font(bold=True, size=11, color="FFFFFF")
h2f = Font(bold=True, size=12, color="1F3864")
hfill = PatternFill("solid", fgColor="2F5496")
secfill = PatternFill("solid", fgColor="D6E4F0")
done_f = PatternFill("solid", fgColor="C6EFCE")
part_f = PatternFill("solid", fgColor="FFEB9C")
fail_f = PatternFill("solid", fgColor="FFC7CE")
info_f = PatternFill("solid", fgColor="DAEEF3")
wrap = Alignment(wrap_text=True, vertical="top")
bdr = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

def setup_headers(ws, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hf; c.fill = hfill; c.border = bdr
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w

def add_rows(ws, rows, start_row=2, status_col=None):
    for r, rd in enumerate(rows, start_row):
        is_sec = rd[0] and str(rd[0]).startswith(">>")
        for ci, val in enumerate(rd, 1):
            if is_sec and ci == 1:
                val = str(val).replace(">>", "").strip()
            c = ws.cell(row=r, column=ci, value=val)
            c.alignment = wrap; c.border = bdr
            if is_sec:
                c.fill = secfill; c.font = Font(bold=True, size=11)
            elif status_col and ci == status_col:
                s = str(val) if val else ""
                if "已完成" in s or "可用" in s: c.fill = done_f
                elif "部分" in s or "待优化" in s or "需配置" in s: c.fill = part_f
                elif "未完成" in s or "Mock" in s or "未开发" in s: c.fill = fail_f

# ══════════════════════════════════════════════════
# Sheet 1: OZON 业务流程总览
# ══════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "OZON业务流程总览"
setup_headers(ws1, ["阶段","步骤","说明","输入","输出","负责模块","状态"], [6,18,40,25,25,20,10])
add_rows(ws1, [
    [">>阶段一：商品采集（数据入口）","","","","","",""],
    ["1","浏览器插件采集","Chrome扩展在1688/淘宝/天猫商品页点击采集按钮，自动提取DOM数据","商品详情页URL","OzonSource+SKU+Media","browser-extension/content.js\nozon.py api_source_add()","已完成"],
    ["1","上传1688 HTML","用SingleFile扩展保存1688页面为HTML，上传到系统自动解析","SingleFile HTML文件","OzonSource+SKU+Media","ozon.py source_import_1688_html()\njiyun/1688 AlibabaParser","已完成"],
    ["1","导入1688文件夹","先用jiyun/1688工具处理HTML生成图片文件夹，再导入","jiyun/1688输出文件夹","OzonSource+SKU+Media","ozon.py source_import_1688_folder()","已完成"],
    ["1","网页URL采集","粘贴商品URL，后端自动抓取+AI解析提取结构化数据","商品URL","OzonSource+SKU+Media","ozon.py source_collect_url()\nozon_collector.py","已完成"],
    ["1","手动粘贴JSON","直接粘贴符合格式的JSON商品数据","JSON文本","OzonSource","ozon.py source_add()","已完成"],

    [">>阶段二：商品适配（数据标准化）","","","","","",""],
    ["2","创建适配组","将一个或多个源商品关联为一个适配任务","OzonSource","SourceProductGroup","ozon.py adaptation_workspace()","已完成"],
    ["2","标准化事实库","从原始数据中提取标准商品名、品牌、型号、材质等","采集的原始数据","ProductFact+ProductFactSku","ozon.py api_save_fact()","已完成"],
    ["2","AI智能填充","AI根据已有信息自动补全缺失的商品属性","ProductFact","补全的属性字段","ozon.py api_ai_suggest()","需配置API Key"],
    ["2","AI图片识别","视觉模型从商品图片中提取品牌、颜色、材质等信息","OzonSourceMedia图片","ImageFact","ozon.py api_analyze_images()","需配置Vision模型"],
    ["2","事实审核","人工确认标准化后的商品信息是否准确","ProductFact","审核通过状态","ozon.py api_approve_fact()","已完成"],

    [">>阶段三：OZON类目映射","","","","","",""],
    ["3","类目树同步","从OZON API拉取完整商品类目树(约5000+类目)","OZON API","OzonCategory","ozon.py api_sync_category_tree()","已完成"],
    ["3","AI类目推荐","根据商品标题和属性自动推荐最合适的OZON类目","商品标题+属性","推荐类目列表","ozon.py api_adaptation_recommend_category()","需配置API Key"],
    ["3","类型选择","在类目下选择具体商品类型(如'消费级无人机')","OzonCategory","OzonCategoryType","ozon.py api_adaptation_save_category_type()","已完成"],
    ["3","属性Schema同步","同步该类型下OZON要求的所有属性(必填/选填)","OzonCategoryType","OzonCategoryAttribute","ozon.py api_sync_category_attributes()","已完成"],
    ["3","属性映射","将本地商品字段映射到OZON属性(如材质→Material)","ProductFact字段","OzonAttributeMapping","ozon.py api_save_mapping()","已完成"],
    ["3","间隙分析","检测草稿中缺少哪些OZON必填属性","OzonDraft vs Schema","OzonFieldGap列表","ozon.py api_field_gaps()","已完成"],

    [">>阶段四：AI内容加工","","","","","",""],
    ["4","AI生成俄语标题","根据中文标题+属性+品牌生成俄语商品名称(<=150字符)","中文标题+属性","俄语标题+置信度","ozon.py processing_generate()","Mock模拟,未完成"],
    ["4","AI生成俄语卖点","根据商品特性+图片识别结果生成3-5条俄语卖点","商品属性+图片","俄语卖点列表","ozon.py processing_generate()","Mock模拟,未完成"],
    ["4","AI生成俄语描述","根据规格参数+详情图内容+使用场景本土化编写俄语描述","全部商品数据","俄语HTML描述","ozon.py processing_generate()","Mock模拟,未完成"],

    [">>阶段五：图片方案","","","","","",""],
    ["5","图片槽位分配","为OZON商品安排8个图片槽位(白底主图/场景图/功能图等)","OzonSourceMedia","OzonImageSlot(8槽)","ozon.py image_plan()","已完成"],
    ["5","图片审核","逐张审核图片是否适合OZON上架","OzonImageSlot","approved/rejected","ozon.py image_plan_approve()","已完成"],
    ["5","图片增删","删除不合适的图片，补充新图片","人工操作","更新Media列表","processing.html","未完成"],

    [">>阶段六：刊登发布","","","","","",""],
    ["6","生成草稿","从适配方案生成OZON listing草稿","ListingAdaptation","OzonDraft+DraftSku","ozon.py api_generate_draft()","已完成"],
    ["6","草稿验证","检查草稿数据完整性(标题/描述/类目/价格/库存/属性/图片)","OzonDraft","验证结果","ozon.py listing_validate()","已完成"],
    ["6","草稿审核","人工审核通过后标记为可发布状态","OzonDraft","approved状态","ozon.py listing_approve()","已完成"],
    ["6","发布到OZON","调用OZON Seller API上传商品数据","OzonDraft+Account","OzonPublishJob","ozon.py listing_publish()","需配置OZON账户"],
    ["6","发布任务追踪","查看发布历史，失败任务支持重试","OzonPublishJob","发布状态","ozon.py publish_jobs()","已完成"],

    [">>阶段七：在线商品管理","","","","","",""],
    ["7","同步在线商品","从OZON API拉取当前店铺的所有在线商品","OZON API","OzonOnlineProduct","ozon.py online_products_sync()","需配置OZON账户"],
    ["7","更新价格/库存","批量修改在线商品的价格和库存","新价格/库存","API更新","ozon.py online_product_update()","需配置OZON账户"],
    ["7","上架/下架","Archive/Unarchive商品","操作指令","商品状态变更","ozon.py online_product_archive()","需配置OZON账户"],
], status_col=7)

# ══════════════════════════════════════════════════
# Sheet 2: 采集方式对比
# ══════════════════════════════════════════════════
ws2 = wb.create_sheet("采集方式对比")
setup_headers(ws2, ["采集方式","支持平台","操作步骤","优势","劣势","可靠性","推荐场景"], [16,12,35,30,30,8,20])
add_rows(ws2, [
    ["浏览器插件采集","淘宝\n天猫\n1688\n拼多多",
     "1.安装Chrome插件\n2.打开商品页\n3.点击'采集'按钮\n4.自动发送到后端",
     "一键操作\n实时采集\n无需额外工具","1688新版页面选择器不稳定\n受反爬限制\n详情图采集困难","淘宝天猫:高\n1688:中","淘宝/天猫商品\n快速批量采集"],
    ["上传1688 HTML\n(推荐)","1688",
     "1.安装SingleFile扩展\n2.打开1688商品页\n3.点击SingleFile保存HTML\n4.在系统上传HTML文件",
     "最可靠\n完整数据\n绕过反爬\n自动下载图片","需两步操作\n需安装SingleFile","高","1688商品\n（首选方案）"],
    ["导入1688文件夹","1688",
     "1.SingleFile保存HTML\n2.拖到jiyun/1688工具\n3.在系统输入文件夹路径",
     "图片已下载到本地\n分类清晰(T_/color_/C_)","操作步骤多\n需安装jiyun工具","高","已有jiyun工具的用户"],
    ["网页URL采集","通用",
     "1.复制商品URL\n2.在系统粘贴URL\n3.自动抓取+AI解析",
     "最简单\n支持所有平台","受JS渲染限制\n1688数据可能不全\n需配置AI API Key","中","非1688平台\n页面结构简单的商品"],
    ["手动粘贴JSON","通用",
     "1.准备符合格式的JSON\n2.粘贴到系统",
     "完全可控\n适合开发调试","需手动准备数据","高","开发测试\n特殊数据源"],
])

# ══════════════════════════════════════════════════
# Sheet 3: AI能力矩阵
# ══════════════════════════════════════════════════
ws3 = wb.create_sheet("AI能力矩阵")
setup_headers(ws3, ["AI功能","使用场景","技术方案","输入数据","输出结果","所需配置","当前状态","优先级"], [16,20,18,25,20,18,10,6])
add_rows(ws3, [
    ["AI内容生成","加工页生成俄语标题/卖点/描述","OpenAI GPT-4o\n或 DeepSeek","中文标题+SKU+属性+图片识别结果","俄语标题(<=150字符)\n俄语卖点(3-5条)\n俄语HTML描述","OpenAI或DeepSeek\nAPI Key","Mock模拟","P0"],
    ["AI图片识别","从商品图片提取品牌/颜色/材质/型号","Vision API\nGPT-4o-vision\nQwen-VL","商品主图+详情图","ImageFact列表\n(字段名/值/置信度)","VisionModelConfig\n(provider+API Key)","需配置","P1"],
    ["AI智能填充","自动补全事实库缺失字段","OpenAI/DeepSeek","已有ProductFact字段","补全的属性值","API Key","需配置","P1"],
    ["AI类目推荐","根据商品信息推荐OZON类目","OpenAI/DeepSeek","商品标题+品牌+类目","OZON类目ID+名称\n+匹配理由","API Key","需配置","P1"],
    ["AI类目翻译","翻译OZON类目名俄→中","OpenAI/DeepSeek","俄语类目名列表","中文翻译","API Key","需配置","P2"],
    ["AI商品提取","从URL/HTML提取结构化商品数据","OpenAI GPT-4o","抓取的HTML文本","标题/SKU/价格/规格","API Key","可用","已完成"],
    ["AI仓库助手","对话式库存/订单查询","OpenAI/DeepSeek\n+Tool Calling","用户自然语言问题","结构化查询结果","API Key","可用","已完成"],
], status_col=7)

# ══════════════════════════════════════════════════
# Sheet 4: 图片处理流程
# ══════════════════════════════════════════════════
ws4 = wb.create_sheet("图片处理流程")
setup_headers(ws4, ["环节","说明","图片分类","过滤规则","当前状态","待改进"], [14,35,12,40,10,35])
add_rows(ws4, [
    [">>采集阶段","","","","",""],
    ["插件端分类","浏览器插件按DOM区域将图片分为main/sku/detail/reject","main\nsku\ndetail\nreject",
     "URL关键词过滤(35个)\n中文文本过滤(30+个)\nCDN白名单\n尺寸过滤\n缩略图后缀过滤","已完成","1688新版DOM选择器需持续适配"],
    ["后端分类","classify_source_image_url()级联分类器,8级过滤","usable\nneeds_review\nrejected",
     "Level0:尺寸\nLevel1:扩展名\nLevel2:域名\nLevel3:文本关键词\nLevel4:区域规则\nLevel5:URL路径\nLevel6:URL模式\nLevel7:文本检查","已完成",""],
    ["HTML导入分类","jiyun/1688按文件名前缀自动分类","T_→main\ncolor_→sku\nC_→detail",
     "文件名模式匹配","已完成",""],

    [">>适配阶段","","","","",""],
    ["图库展示","适配工作台5组分类展示所有图片","主图\nSKU图\n详情图\n待审核\n已过滤",
     "按compliance_status和role分组","已完成",""],
    ["SKU图绑定","每张SKU图显示绑定的SKU名称","linked_sku_name",
     "从采集时绑定或手动关联","已完成",""],
    ["人工补图","支持上传文件或粘贴URL补充图片","用户指定role",
     "无自动分类","已完成",""],
    ["图片恢复","被过滤的图片可人工恢复为可用","rejected→usable",
     "人工判断","已完成",""],

    [">>发布阶段","","","","",""],
    ["图片方案","8槽位图片安排(OZON要求)","槽位1-8",
     "按用途分配:\n1白底主图\n2-3场景图\n4-5功能图\n6尺寸图\n7包装图\n8品牌图","已完成","AI图片生成未开发"],
    ["图片审核","逐张审核是否适合上架","approved/rejected",
     "人工审核","已完成",""],
    ["图片增删","在加工页增加/删除图片","","","未完成","需开发加工页图片管理UI"],
], status_col=5)

# ══════════════════════════════════════════════════
# Sheet 5: 定价体系
# ══════════════════════════════════════════════════
ws5 = wb.create_sheet("定价体系")
setup_headers(ws5, ["参数","说明","数据来源","示例值","配置位置"], [14,35,20,12,22])
add_rows(ws5, [
    ["采购价(CNY)","源商品的采购单价(人民币)","采集/手动填写","31.00","OzonSourceSku.purchase_price_cny"],
    ["汇率","CNY→RUB实时汇率","系统自动更新(每小时)","12.5","ExchangeRate表"],
    ["利润率","期望的利润百分比","手动配置","30%","OzonPricingRule.margin"],
    ["OZON佣金","OZON平台扣除的佣金比例(按类目不同)","手动配置","15%","OzonPricingRule.commission"],
    ["物流费","从中国到俄罗斯的物流费用","手动配置/按重量","50 RUB","OzonPricingRule.logistics"],
    ["","","","",""],
    ["定价公式","OZON售价 = (采购价 x 汇率 + 物流费) x (1+利润率) / (1-佣金率)","","",""],
    ["计算示例","(31 x 12.5 + 50) x 1.30 / 0.85 = 653 RUB","","653 RUB",""],
])

# ══════════════════════════════════════════════════
# Sheet 6: 数据模型关系
# ══════════════════════════════════════════════════
ws6 = wb.create_sheet("OZON数据模型")
setup_headers(ws6, ["模型","中文名","分类","关键字段","关联关系","说明"], [22,14,10,40,30,30])
add_rows(ws6, [
    ["OzonAccount","OZON店铺","核心","name, client_id, api_key, is_active","User(FK)","OZON Seller API认证信息"],
    ["OzonSource","采集来源","核心","platform, title_cn, source_url, status,\nraw_json, quality_json, detail_missing","User(FK)","采集的源商品数据,支持软删除"],
    ["OzonSourceSku","来源SKU","核心","source_sku_name, color_cn, style_cn,\npurchase_price_cny","OzonSource(FK)","源商品的每个规格变体"],
    ["OzonSourceMedia","来源图片","核心","role(main/sku/detail), source_url,\nlocal_path, compliance_status,\nreject_reason, raw_json","OzonSource(FK)","源商品图片,含分类和过滤状态"],
    ["","","","","",""],
    ["SourceProductGroup","适配组","适配层","name, relation_type(1:1/1:N/N:1)","User(FK)","将源商品组织为适配任务"],
    ["SourceProductGroupItem","组-源映射","适配层","","Group(FK), Source(FK)","多对多关联"],
    ["ProductFact","商品事实","适配层","standard_name, category, brand,\nmodel, material, origin, warranty","Group(FK)","标准化的商品信息(中间层)"],
    ["ProductFactSku","事实SKU","适配层","sku_name, color, size, style, price","ProductFact(FK)","标准化的SKU信息"],
    ["ProductFactEvidence","事实证据","适配层","field_path, value, confidence","ProductFact(FK),\nOzonSourceMedia(FK)","每个字段值的来源证据"],
    ["ListingAdaptation","适配方案","适配层","category_id, type_id, status","Group(FK), ProductFact(FK)","事实→草稿的映射方案"],
    ["","","","","",""],
    ["OzonDraft","刊登草稿","发布","title_ru, description_ru, bullets_ru,\ncategory, status, ai_confidence","OzonSource(FK)","待发布的OZON商品草稿"],
    ["OzonDraftSku","草稿SKU","发布","sku_name_ru, price_rub, stock,\nbarcode, ozon_sku_id","OzonDraft(FK)","草稿的每个SKU"],
    ["OzonImageSlot","图片槽位","发布","slot_order(1-8), role, scope,\nsource_url, local_path, status","OzonDraft(FK)","8个图片位置的安排"],
    ["OzonPublishJob","发布任务","发布","status, ozon_product_id, error_message","OzonDraft(FK),\nOzonAccount(FK)","发布到OZON的任务记录"],
    ["","","","","",""],
    ["OzonCategory","OZON类目","类目","category_id, name_ru, name_cn,\nparent_id, level, is_leaf","","OZON类目树(约5000+节点)"],
    ["OzonCategoryType","类目类型","类目","type_id, name_ru, name_cn","OzonCategory(FK)","类目下的商品类型(叶子级)"],
    ["OzonCategoryAttribute","类目属性","类目","attr_id, name, required, type,\ndescription","OzonCategoryType(FK)","该类型要求的商品属性Schema"],
    ["OzonAttributeValue","属性字典值","类目","value_id, value_ru, value_cn","OzonCategoryAttribute(FK)","属性的可选值(如品牌列表)"],
    ["OzonAttributeMapping","属性映射","类目","local_field, transform_rule","OzonCategoryType(FK)","本地字段→OZON属性的映射规则"],
    ["OzonFieldGap","字段缺口","类目","attr_id, status, suggested_value","ListingAdaptation(FK)","缺失的必填属性"],
    ["","","","","",""],
    ["VisionModelConfig","视觉模型配置","AI","provider, model_name, api_key, enabled","User(FK)","Vision API配置"],
    ["ImageAnalysisJob","图片分析任务","AI","task_type, status, response_json","OzonSourceMedia(FK)","图片识别任务记录"],
    ["ImageFact","图片识别事实","AI","field_path, value, confidence, accepted","OzonSourceMedia(FK)","从图片中提取的商品属性"],
    ["","","","","",""],
    ["OzonOnlineProduct","在线商品","运营","product_id, name, status, price, stock","OzonAccount(FK)","OZON上的在线商品缓存"],
    ["OzonOnlineProductAction","操作日志","运营","action, status, details","OzonOnlineProduct(FK)","在线商品操作审计"],
    ["OzonPrompt","提示词模板","配置","name, type(title/bullets/desc/image),\ncontent","User(FK)","AI Prompt模板"],
    ["OzonPricingRule","定价规则","配置","name, margin, commission, logistics,\nexchange_rate_source","User(FK)","定价计算规则"],
])

# ══════════════════════════════════════════════════
# Sheet 7: 待办事项
# ══════════════════════════════════════════════════
ws7 = wb.create_sheet("待办事项")
setup_headers(ws7, ["优先级","任务","模块","说明","预估工作量","依赖"], [6,28,16,45,10,20])
add_rows(ws7, [
    [">>P0 — 必须完成（核心功能缺失）","","","","",""],
    ["P0","AI内容生成:Mock转真实API","商品加工","当前processing_generate()返回Mock数据,需改为调用OpenAI/DeepSeek生成真实俄语标题/卖点/描述","4小时","API Key配置"],
    ["P0","AI Prompt优化:本土化内容","商品加工","AI应基于:商品属性+SKU信息+图片识别结果+采集内容 生成本土化俄语描述,而非简单翻译","3小时","AI内容生成完成"],
    ["P0","加工页图片增删功能","商品加工","在加工页支持:删除不需要的图片/添加新图片/调整图片角色(主图→SKU→详情)","3小时","无"],

    [">>P1 — 重要改进","","","","",""],
    ["P1","1688插件采集稳定性","商品采集","window.viewData JSON提取待验证;DOM选择器需适配更多1688页面版本","2小时","测试页面"],
    ["P1","图片后端代理下载","图片管理","采集时直接在后端下载图片到本地,避免前端referrer防盗链问题","2小时","无"],
    ["P1","适配页图片加载优化","商品适配","部分alicdn图片因referrer限制加载失败,已加referrerpolicy但不够彻底","1小时","图片代理"],
    ["P1","完善定价规则应用","定价体系","将定价规则自动应用到草稿SKU的价格计算","2小时","无"],

    [">>P2 — 体验优化","","","","",""],
    ["P2","多1688页面版本兼容","商品采集","测试服装/电子/食品等不同类型1688页面确保选择器覆盖率","3小时","测试页面"],
    ["P2","图片质量自动检测","图片管理","自动检测服务图标/资质认证/营业执照等非商品图片","2小时","无"],
    ["P2","批量操作优化","全局","采集列表批量导入/加工页批量处理/草稿批量发布","4小时","无"],
    ["P2","AI图片生成(OZON白底图)","图片方案","用AI生成白底主图/场景图等OZON要求的图片","8小时","AI API"],
    ["P2","调试日志清理","全局","移除content.js中的[1688-DEBUG]调试日志,改为可配置开关","1小时","无"],
], status_col=1)
# Apply priority colors
for row in ws7.iter_rows(min_row=2, max_row=ws7.max_row, min_col=1, max_col=1):
    for cell in row:
        v = str(cell.value) if cell.value else ""
        if v == "P0": cell.fill = fail_f
        elif v == "P1": cell.fill = part_f
        elif v == "P2": cell.fill = done_f

path = os.path.join("G:", os.sep, "inventory", "docs", "OZON_detailed_guide_20260620.xlsx")
wb.save(path)
print("OK:", path)
