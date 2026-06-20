"""生成完整系统功能清单 Excel"""
import openpyxl, os
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
hf = Font(bold=True, size=11, color="FFFFFF")
hfill = PatternFill("solid", fgColor="2F5496")
secfill = PatternFill("solid", fgColor="D6E4F0")
done_f = PatternFill("solid", fgColor="C6EFCE")
part_f = PatternFill("solid", fgColor="FFEB9C")
fail_f = PatternFill("solid", fgColor="FFC7CE")
wrap = Alignment(wrap_text=True, vertical="top")
bdr = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

def make_sheet(ws, headers, widths, rows, status_col=None, section_col=None):
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hf; c.fill = hfill; c.border = bdr
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    for r, rd in enumerate(rows, 2):
        is_section = rd[0] and str(rd[0]).startswith("==")
        for ci, val in enumerate(rd, 1):
            if is_section and ci == 1:
                val = str(val).replace("==", "").strip()
            c = ws.cell(row=r, column=ci, value=val)
            c.alignment = wrap; c.border = bdr
            if is_section:
                c.fill = secfill; c.font = Font(bold=True, size=11)
            elif status_col and ci == status_col:
                s = str(val) if val else ""
                if "已完成" in s or "正常" in s: c.fill = done_f
                elif "部分" in s or "待优化" in s: c.fill = part_f
                elif "未完成" in s or "未开发" in s or "模拟" in s: c.fill = fail_f

# ═══ Sheet 1: 全系统功能清单 ═══
ws1 = wb.active
ws1.title = "全系统功能清单"
H1 = ["序号","功能模块","页面/功能","路由","完成状态","功能说明","存在的问题"]
W1 = [6, 14, 24, 35, 10, 40, 40]
n = 0
def row(mod, page, route, status, desc, issue=""):
    global n; n += 1
    return [n, mod, page, route, status, desc, issue]
def sec(title):
    return ["==" + title, "", "", "", "", "", ""]

R1 = [
    sec("一、用户认证"),
    row("用户认证","注册页","/register","已完成","用户注册(限3次/小时)",""),
    row("用户认证","登录页","/login","已完成","用户登录(限5次/分钟)",""),
    row("用户认证","退出","/logout","已完成","退出登录",""),

    sec("二、首页仓库概览"),
    row("仓库概览","首页仪表盘","/","已完成","对账时段汇总/供应商订单/客户订单/库存概览/负库存预警/仓库供需对比",""),

    sec("三、基础资料"),
    row("产品管理","产品列表","/products","已完成","产品CRUD/搜索/分页/SKU自动生成",""),
    row("产品管理","套装/拆组","/products/bundles","已完成","套装组合方案管理",""),
    row("客户管理","客户列表","/customers","已完成","客户CRUD/搜索/分页",""),
    row("供应商管理","供应商列表","/suppliers","已完成","供应商CRUD/搜索/分页",""),

    sec("四、供应商订单 + 入库"),
    row("供应商订单","订单列表","/supplier_orders","已完成","供应商订单列表/对账时段拆分",""),
    row("供应商订单","新增订单","/supplier_orders/add","已完成","创建供应商订单/自动编号",""),
    row("供应商订单","编辑订单","/supplier_orders/edit/<id>","已完成","编辑(有收货则锁定)",""),
    row("供应商订单","收货表单","/supplier_orders/receive/<id>","已完成","显示待收数量/创建入库单",""),
    row("入库管理","入库单列表","/receipts","已完成","入库单列表含明细",""),
    row("入库管理","新增入库","/purchase/add","已完成","创建入库单/自动触发拆分订单",""),
    row("入库管理","编辑入库","/receipts/edit/<id>","已完成","编辑入库单",""),

    sec("五、客户订单 + 出库"),
    row("客户订单","订单列表","/orders","已完成","客户订单列表/对账时段/自动截单/客户筛选",""),
    row("客户订单","新增订单","/orders/add","已完成","创建客户订单",""),
    row("客户订单","发货表单","/orders/ship/<id>","已完成","显示待发数量/支持截单/替代商品转新订单",""),
    row("客户订单","修正数量","/orders/fix/<id>","已完成","按实发调减订单量",""),
    row("客户订单","结算订单","/orders/settle/<id>","已完成","零价出库剩余商品",""),
    row("出库管理","出库单列表","/shipments","已完成","出库单列表",""),
    row("出库管理","新增出库","/sales/add","已完成","创建出库单/库存校验",""),

    sec("六、库存工具"),
    row("拆分规则","拆分规则列表","/inventory/split-rules","已完成","产品拆分(拆机)规则管理",""),
    row("拆分订单","拆分订单列表","/inventory/split-orders","已完成","拆分订单(草稿/确认/取消)/库存联动",""),
    row("组装规则","组装规则列表","/inventory/assembly-rules","已完成","产品组装(组合)规则管理",""),
    row("组装订单","组装订单列表","/inventory/assembly-orders","已完成","组装订单(草稿/确认/取消)/库存联动",""),

    sec("七、财务管理"),
    row("客户财务","客户财务概览","/customer/finance","已完成","订单总额/已发货/退款/应收/余额",""),
    row("退款管理","退款列表","/refunds","已完成","退款记录CRUD/计划退款设置",""),

    sec("八、统计报表"),
    row("报表","每日进出货","/report/daily","已完成","日进出统计图表(数量+金额/产品明细)",""),
    row("报表","客户订单报表","/report/customer","已完成","客户订单履行报表/对账时段",""),
    row("报表","供应商订单报表","/report/supplier","已完成","供应商到货报表/对账时段",""),
    row("报表","库存报表","/report/inventory","已完成","当前库存/均价/库存货值/负库存预警",""),
    row("报表","销售利润报表","/report/sales_profit","已完成","收入/成本/毛利(按产品)",""),
    row("报表","库存趋势图","/report/inventory_trend","已完成","库存量时间趋势图",""),
    row("报表","周期进出货汇总","/report/inventory_period","已完成","期间进出/金额/利润/利润率",""),

    sec("九、工具箱"),
    row("汇率工具","汇率换算","/exchange","已完成","CNY/RUB/USD自动更新(每小时)/换算计算器",""),
    row("定价工具","定价计算器","/tools/pricing","已完成","独立定价计算工具",""),
    row("数据导入导出","数据管理","/data","已完成","CSV导入导出(产品/客户/供应商)",""),
    row("AI导入","AI智能导入","/ai-import","已完成","上传PDF/Excel由AI提取产品信息","需配置API Key"),
    row("AI助手","AI仓库助手","/ai-assistant","已完成","对话式AI(支持库存查询/订单查询/文件分析)","需配置API Key"),

    sec("十、操作日志"),
    row("操作日志","操作日志查看","/admin/logs","已完成","管理员审计日志(分页)",""),

    sec("十一、OZON运营 — 商品采集"),
    row("OZON工作台","运营仪表盘","/ozon/dashboard","已完成","状态计数概览",""),
    row("店铺管理","OZON账户管理","/ozon/accounts","已完成","OZON API账户CRUD/连接测试/插件Token",""),
    row("商品采集","采集列表页","/ozon/sources","已完成","采集商品列表/回收站/筛选",""),
    row("商品采集","浏览器插件采集(淘宝/天猫)","/ozon/api/sources/add","已完成","插件一键采集标题/SKU/主图(不采详情图)",""),
    row("商品采集","浏览器插件采集(1688)","/ozon/api/sources/add","部分完成","5个专用函数+viewData JSON提取","新版1688 DOM选择器不稳定"),
    row("商品采集","上传1688 HTML","/ozon/sources/import-1688-html","已完成","[新增]集成jiyun/1688解析器自动提取","需安装SingleFile扩展"),
    row("商品采集","导入1688文件夹","/ozon/sources/import-1688-folder","已完成","[新增]读取jiyun/1688输出文件夹","需先运行jiyun工具"),
    row("商品采集","网页URL采集","/ozon/sources/collect-url","已完成","URL粘贴+AI解析+headless兜底","1688 JS渲染可能不全"),
    row("商品采集","手动粘贴JSON","/ozon/sources/add","已完成","手动粘贴采集JSON数据",""),
    row("商品采集","商品详情页","/ozon/sources/<id>","已完成","源商品数据查看/图片预览",""),

    sec("十二、OZON运营 — 商品加工"),
    row("商品加工","商品加工页(AI生成)","/ozon/processing/<id>","部分完成","源数据展示+AI俄语内容生成","1.AI为Mock模拟\n2.图片无增删\n3.内容不基于商品数据本土化"),
    row("商品适配","适配工作台","/ozon/adaptation/<id>","部分完成","三栏:源数据/事实库/OZON草稿;图库5组;SKU绑定;人工补图","部分图片加载失败"),
    row("事实库","事实库列表","/ozon/fact-library","已完成","标准化商品事实浏览/审核",""),
    row("AI内容","AI加工生成","/ozon/processing/<id>/generate","未完成","Mock模拟数据,未调用真实AI","需重写为真实API调用"),
    row("AI内容","AI图片识别","/ozon/api/adaptation/<id>/analyze-images","部分完成","视觉模型图片分析","需配置VisionModelConfig"),
    row("AI内容","AI智能填充","/ozon/api/adaptation/<id>/ai-suggest","部分完成","AI属性建议","需配置API Key"),

    sec("十三、OZON运营 — 图片管理"),
    row("图片管理","图片过滤/恢复","api/source-media/<id>/restore","已完成","恢复被过滤的图片",""),
    row("图片管理","人工补图(上传/URL)","api/source-media/<id>/upload","已完成","[新增]文件上传+URL粘贴",""),
    row("图片管理","图片文件服务","/ozon/uploads/source_media/...","已完成","[新增]本地图片访问路由",""),
    row("图片管理","图片方案","/ozon/image-plan/<draft_id>","已完成","8槽图片方案管理/审批",""),
    row("图片管理","图片分类系统","classify_source_image_url()","已完成","main/sku/detail/reject四类",""),

    sec("十四、OZON运营 — 刊登发布"),
    row("刊登管理","草稿列表","/ozon/listings","已完成","草稿列表/筛选/批量删除",""),
    row("刊登管理","草稿审核","/ozon/listings/<id>","已完成","审核页/验证/批准",""),
    row("刊登管理","发布到OZON","/ozon/listings/<id>/publish","部分完成","调用OZON API发布","需配置OZON API账户"),
    row("发布任务","发布任务列表","/ozon/publish-jobs","已完成","发布历史/重试失败任务",""),

    sec("十五、OZON运营 — 类目属性"),
    row("类目管理","类目属性浏览","/ozon/category-attributes","已完成","OZON类目树/属性Schema/字典值",""),
    row("类目管理","类目同步","/ozon/api/category/sync-tree","已完成","从OZON API同步类目树",""),
    row("类目管理","类目翻译","/ozon/api/category/translate","已完成","AI翻译类目名(俄->中)",""),
    row("类目管理","属性同步","/ozon/api/category/<id>/sync-attributes","已完成","同步属性Schema+字典值",""),
    row("类目管理","间隙分析","/ozon/api/category/<id>/gaps/<draft_id>","已完成","缺失必填字段检测",""),

    sec("十六、OZON运营 — 其他"),
    row("提示词库","提示词管理","/ozon/prompts","已完成","AI Prompt模板(标题/卖点/描述/图片)",""),
    row("定价规则","定价规则管理","/ozon/pricing","已完成","利润率/佣金/物流/汇率规则",""),
    row("视觉模型","模型配置","/ozon/models","已完成","Vision API配置(OpenAI/Qwen/Gemini)",""),
    row("在线商品","在线商品管理","/ozon/online-products","已完成","同步/浏览/上下架OZON在线商品","需配置OZON账户"),
]

make_sheet(ws1, H1, W1, R1, status_col=5)

# ═══ Sheet 2: 数据模型 ═══
ws2 = wb.create_sheet("数据模型(56张表)")
H2 = ["序号","分类","模型名称","说明","关键字段"]
W2 = [5, 14, 22, 30, 50]
models = [
    [1,"核心","User","用户账户","username, password_hash, is_admin, extension_token"],
    [2,"核心","Product","产品","name, sku, brand, category, unit_cost, stock(缓存)"],
    [3,"核心","ProductBundle","套装方案","name, product(主产品)"],
    [4,"核心","ProductBundleItem","套装明细","bundle, product, quantity"],
    [5,"核心","Customer","客户","name, contact, planned_refund"],
    [6,"核心","Supplier","供应商","name, contact"],
    [7,"库存转换","ProductSplitRule","拆分规则","source_product, name"],
    [8,"库存转换","ProductSplitRuleItem","拆分明细","rule, target_product, quantity, cost_ratio"],
    [9,"库存转换","ProductSplitOrder","拆分订单","rule, status(draft/confirmed/cancelled)"],
    [10,"库存转换","ProductSplitOrderItem","拆分订单明细","order, product, quantity, unit_cost"],
    [11,"库存转换","ProductAssemblyRule","组装规则","target_product, name"],
    [12,"库存转换","ProductAssemblyRuleItem","组装明细","rule, component_product, quantity"],
    [13,"库存转换","ProductAssemblyOrder","组装订单","rule, status"],
    [14,"库存转换","ProductAssemblyOrderItem","组装订单明细","order, product, quantity"],
    [15,"订单","SupplierOrder","供应商订单","supplier, order_number, status, total"],
    [16,"订单","SupplierOrderItem","供应商订单明细","order, product, quantity, price"],
    [17,"订单","PurchaseOrder","入库单","supplier_order, date, total"],
    [18,"订单","PurchaseOrderItem","入库单明细","order, product, quantity, price"],
    [19,"订单","CustomerOrder","客户订单","customer, order_number, status, total"],
    [20,"订单","CustomerOrderItem","客户订单明细","order, product, quantity, price"],
    [21,"订单","SalesOrder","出库单","customer_order, date, tracking_number"],
    [22,"订单","SalesOrderItem","出库单明细","order, product, quantity, price"],
    [23,"财务","CustomerRefund","退款记录","customer, amount, reason, date"],
    [24,"财务","CustomerTransaction","客户交易","customer, type, amount, balance"],
    [25,"系统","ExchangeRate","汇率","currency_pair, rate, updated_at"],
    [26,"系统","OperationLog","操作日志","user, action_type, target, details"],
    [27,"系统","UserApiKey","AI密钥","user, provider, encrypted_key"],
    [28,"OZON核心","OzonAccount","OZON店铺","name, client_id, api_key"],
    [29,"OZON核心","OzonSource","采集来源","platform, title_cn, source_url, status, raw_json"],
    [30,"OZON核心","OzonSourceSku","来源SKU","source, sku_name, color_cn, price"],
    [31,"OZON核心","OzonSourceMedia","来源图片","source, role, source_url, local_path, compliance_status, raw_json"],
    [32,"OZON核心","OzonDraft","刊登草稿","source, title_ru, description_ru, category, status"],
    [33,"OZON核心","OzonDraftSku","草稿SKU","draft, sku_name, price_rub, stock"],
    [34,"OZON核心","OzonImageSlot","图片槽位","draft, slot_order, role, source_url"],
    [35,"OZON核心","OzonPublishJob","发布任务","draft, account, status, ozon_product_id"],
    [36,"OZON核心","OzonPrompt","提示词模板","name, type, content"],
    [37,"OZON核心","OzonPricingRule","定价规则","name, margin, commission, logistics"],
    [38,"OZON适配","SourceProductGroup","适配组","name, relation_type"],
    [39,"OZON适配","SourceProductGroupItem","组-源映射","group, source"],
    [40,"OZON适配","ProductFact","商品事实","group, standard_name, category, brand"],
    [41,"OZON适配","ProductFactSku","事实SKU","fact, sku_name, color, size, price"],
    [42,"OZON适配","ProductFactEvidence","事实证据","fact, field_path, value, confidence"],
    [43,"OZON适配","ListingAdaptation","适配方案","group, fact, category_id, type_id, status"],
    [44,"OZON类目","OzonCategory","OZON类目","category_id, name_ru, name_cn, parent_id"],
    [45,"OZON类目","OzonCategoryType","类目类型","type_id, category, name_ru, name_cn"],
    [46,"OZON类目","OzonCategoryAttribute","类目属性","category_type, attr_id, name, required, type"],
    [47,"OZON类目","OzonAttributeValue","属性字典值","attribute, value_id, value_ru, value_cn"],
    [48,"OZON类目","OzonAttributeMapping","属性映射","category_type, attr_id, local_field"],
    [49,"OZON类目","OzonFieldGap","字段缺口","adaptation, attr_id, status"],
    [50,"OZON同步","OzonCategorySyncJob","同步任务","category_id, job_type, status, result"],
    [51,"OZON同步","OzonFavoriteCategoryType","收藏类型","user, category_type"],
    [52,"OZON视觉","VisionModelConfig","视觉模型配置","provider, model_name, api_key, enabled"],
    [53,"OZON视觉","ImageAnalysisJob","图片分析任务","media, task_type, status, response"],
    [54,"OZON视觉","ImageFact","图片识别事实","media, field_path, value, confidence, accepted"],
    [55,"OZON在线","OzonOnlineProduct","在线商品","account, product_id, name, status, price"],
    [56,"OZON在线","OzonOnlineProductAction","操作日志","product, action, status"],
]
make_sheet(ws2, H2, W2, models)

# ═══ Sheet 3: 统计 ═══
ws3 = wb.create_sheet("系统统计")
H3 = ["指标","数值"]
W3 = [25, 15]
stats = [
    ["蓝图(Blueprint)数量", "22个"],
    ["路由总数", "约145个"],
    ["数据模型(表)数量", "56张"],
    ["模板文件数量", "56个"],
    ["Python代码行数(估)", "约15,000行"],
    ["JS代码行数(content.js)", "约2,200行"],
    ["",""],
    ["主系统(进销存)路由", "约65个"],
    ["OZON模块路由", "约80个"],
    ["",""],
    ["已完成功能", "约85%"],
    ["部分完成功能", "约10%"],
    ["未完成功能", "约5%"],
    ["",""],
    ["本次会话改动文件", "10个"],
    ["本次会话新增路由", "4个"],
    ["本次会话删除代码", "约500行(H5补采)"],
    ["本次会话新增代码", "约800行"],
]
make_sheet(ws3, H3, W3, stats)

path = os.path.join("G:", os.sep, "inventory", "docs", "system_overview_20260620.xlsx")
os.makedirs(os.path.dirname(path), exist_ok=True)
wb.save(path)
print("OK:", path)
