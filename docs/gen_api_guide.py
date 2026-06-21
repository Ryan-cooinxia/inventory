"""生成 OZON Seller API 使用文档"""
import openpyxl, os
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
hf = Font(bold=True, size=11, color="FFFFFF")
hfill = PatternFill("solid", fgColor="2F5496")
secfill = PatternFill("solid", fgColor="D6E4F0")
done_f = PatternFill("solid", fgColor="C6EFCE")
part_f = PatternFill("solid", fgColor="FFEB9C")
fail_f = PatternFill("solid", fgColor="FFC7CE")
wrap = Alignment(wrap_text=True, vertical="top")
bdr = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

def setup(ws, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hf; c.fill = hfill; c.border = bdr
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w

def rows(ws, data, start=2, status_col=None):
    for r, rd in enumerate(data, start):
        is_sec = rd[0] and str(rd[0]).startswith(">>")
        for ci, val in enumerate(rd, 1):
            if is_sec and ci == 1: val = str(val).replace(">>","").strip()
            c = ws.cell(row=r, column=ci, value=val)
            c.alignment = wrap; c.border = bdr
            if is_sec: c.fill = secfill; c.font = Font(bold=True)
            elif status_col and ci == status_col:
                s = str(val) if val else ""
                if "已实现" in s: c.fill = done_f
                elif "未实现" in s: c.fill = fail_f

# ═══ Sheet 1: API 概览 ═══
ws1 = wb.active
ws1.title = "API概览与认证"
setup(ws1, ["项目","内容"], [20, 70])
rows(ws1, [
    ["API 基础地址", "https://api-seller.ozon.ru"],
    ["官方文档", "https://docs.ozon.ru/api/seller/"],
    ["通信协议", "HTTPS + JSON"],
    ["请求方法", "所有接口均使用 POST 方法"],
    ["",""],
    [">>认证方式",""],
    ["认证方式", "HTTP Header 认证"],
    ["Header: Client-Id", "OZON 卖家后台获取的 Client ID (字符串)"],
    ["Header: Api-Key", "OZON 卖家后台获取的 API Key (字符串)"],
    ["Header: Content-Type", "application/json (固定)"],
    ["",""],
    [">>请求格式",""],
    ["请求体格式", "JSON (application/json)"],
    ["响应格式", "JSON"],
    ["字符编码", "UTF-8"],
    ["超时设置", "30 秒 (REQUEST_TIMEOUT)"],
    ["",""],
    [">>重试与限流",""],
    ["速率限制", "HTTP 429 表示请求过多，需等待后重试"],
    ["重试次数", "最多 2 次重试 (MAX_RETRIES=2)"],
    ["重试间隔", "指数退避: 2秒 x (尝试次数+1)，即 2s, 4s"],
    ["可重试状态码", "429, 500, 502, 503, 504"],
    ["不可重试状态码", "401, 403, 其他 4xx (立即失败)"],
    ["",""],
    [">>错误码说明",""],
    ["401 Unauthorized", "Client-Id 或 Api-Key 无效"],
    ["403 Forbidden", "无权限访问该接口"],
    ["400 Bad Request", "请求参数验证失败 (如缺少必填字段、格式错误)"],
    ["404 Not Found", "请求的资源不存在"],
    ["429 Too Many Requests", "请求过于频繁，触发限流"],
    ["500 Internal Server Error", "OZON 服务器内部错误"],
    ["502 Bad Gateway", "OZON 网关错误"],
    ["503 Service Unavailable", "OZON 服务暂时不可用"],
])

# ═══ Sheet 2: 全部 API 端点 ═══
ws2 = wb.create_sheet("API端点清单")
setup(ws2, ["序号","分类","方法名","HTTP路径","用途","请求体格式","响应格式","状态"], [4,12,26,35,30,40,30,8])
rows(ws2, [
    [">>商品管理","","","","","","",""],
    [1,"商品","import_product()","POST /v3/product/import",
     "创建或更新商品(全量覆盖)",
     '{"items": [{\n  "offer_id": "本地ID",\n  "name": "俄语标题",\n  "category_id": 类目ID,\n  "price": "价格",\n  "description": "描述",\n  "attributes": [...],\n  "images": [...]\n}]}',
     '{"result": {"task_id": 123}}',
     "已实现"],
    [2,"商品","import_product_info()","POST /v1/product/import/info",
     "查询商品导入任务状态",
     '{"task_id": 12345}',
     '{"result": {"status": "...", "items": [...]}}',
     "已实现"],
    [3,"商品","get_product_info()","POST /v3/product/info/list",
     "获取在线商品详细信息",
     '{"offer_id": ["id1","id2"],\n"product_id": [123],\n"sku": [456]}',
     '{"result": {"items": [...]}}',
     "已实现"],
    [4,"商品","list_products()","POST /v3/product/list",
     "查询商品列表(分页)",
     '{"filter": {},\n"last_id": "",\n"limit": 100}',
     '{"result": {"items": [...],\n"total": N, "last_id": "..."}}',
     "已实现"],
    [5,"商品","archive_products()","POST /v1/product/archive",
     "归档(下架)商品",
     '{"product_id": [123, 456]}',
     '{"result": true}',
     "已实现"],
    [6,"商品","unarchive_products()","POST /v1/product/unarchive",
     "取消归档(上架)商品",
     '{"product_id": [123, 456]}',
     '{"result": true}',
     "已实现"],

    [">>价格更新","","","","","","",""],
    [7,"价格","update_product_prices()","POST /v4/product/info/prices",
     "更新商品价格(专用端点,不触发全量覆盖)",
     '{"prices": [{\n  "offer_id": "xxx",\n  "price": "100.00",\n  "old_price": "120.00",\n  "min_price": "90.00",\n  "currency_code": "RUB"\n}]}',
     '{"result": [...]}',
     "已实现"],

    [">>库存更新","","","","","","",""],
    [8,"库存","update_product_stocks()","POST /v2/product/import/stocks",
     "更新商品库存(专用端点)",
     '{"stocks": [{\n  "offer_id": "xxx",\n  "product_id": 12345,\n  "stock": 100,\n  "warehouse_id": 0\n}]}',
     '{"result": [...]}',
     "已实现"],

    [">>类目与属性","","","","","","",""],
    [9,"类目","get_category_tree()","POST /v1/description-category/tree",
     "获取OZON商品类目树(约5000+节点)",
     '{"language": "DEFAULT"}',
     '[{"category_id": "...",\n"title": "...",\n"children": [...]}]',
     "已实现"],
    [10,"类目","get_category_tree_with_subtree()","POST /v1/description-category/tree",
     "获取指定类目的子树(含type_id)",
     '{"language": "DEFAULT",\n"category_id": 12345}',
     '同上 + type_id, type_name',
     "已实现"],
    [11,"类目","get_category_types_for_node()","POST /v1/description-category/tree",
     "获取指定类目下的商品类型列表",
     '{"language": "DEFAULT",\n"category_id": 12345}',
     '{"types": [...],\n"direct_count": N,\n"total_in_tree": N}',
     "已实现"],
    [12,"属性","get_category_attributes()","POST /v1/description-category/attribute",
     "获取指定类型的属性Schema(必填/选填)",
     '{"description_category_id": 123,\n"type_id": 456,\n"language": "DEFAULT"}',
     '[{"attribute_id": ...,\n"name": "...",\n"is_required": true,\n"data_type": "..."}]',
     "已实现"],
    [13,"属性","get_attribute_values()","POST /v1/description-category/attribute/values",
     "获取属性的字典值列表(如品牌列表)",
     '{"description_category_id": 123,\n"type_id": 456,\n"attribute_id": 789,\n"limit": 5000}',
     '[{"id": ..., "value": "..."}]',
     "已实现"],

    [">>图片上传","","","","","","",""],
    [14,"图片","upload_image()","POST /v1/product/pictures/import",
     "通过URL上传商品图片",
     "待确认",
     "待确认",
     "未实现"],

    [">>连接测试","","","","","","",""],
    [15,"测试","test_connectivity()","POST /v3/product/list",
     "测试API连接(查询1个商品验证凭证)",
     '{"filter": {}, "last_id": "", "limit": 1}',
     '{"success": true, "elapsed": 0.5}',
     "已实现"],
], status_col=8)

# ═══ Sheet 3: 关键字段说明 ═══
ws3 = wb.create_sheet("商品数据字段")
setup(ws3, ["字段名","类型","必填","说明","示例值"], [20, 10, 6, 40, 25])
rows(ws3, [
    [">>商品基本字段 (/v3/product/import)","","","",""],
    ["offer_id","string","是","卖家自定义的商品ID(本地标识,唯一)","SKU-001"],
    ["name","string","是","商品俄语标题(最长150字符)","Беспроводной адаптер..."],
    ["category_id","int","是","OZON类目ID(从类目树获取)","17028922"],
    ["type_id","int","否","商品类型ID(从类目类型列表获取)","97311"],
    ["description","string","否","商品俄语描述(支持HTML)","<p>Описание...</p>"],
    ["barcode","string","否","商品条码(EAN/UPC)","4607086560894"],
    ["price","string","否","售价(卢布,字符串格式)","561.50"],
    ["old_price","string","否","原价/划线价(卢布)","699.00"],
    ["min_price","string","否","最低允许售价","500.00"],
    ["vat","string","否","增值税率,默认'0'","0"],
    ["weight","int","否","重量(克)","500"],
    ["width","int","否","宽度(毫米)","150"],
    ["height","int","否","高度(毫米)","100"],
    ["depth","int","否","深度/长度(毫米)","200"],
    ["","","","",""],
    [">>属性(attributes)","","","",""],
    ["attributes","array","否","商品属性数组",""],
    ["  .id","int","是","属性ID(从Schema获取)","85"],
    ["  .values","array","是","属性值数组",""],
    ["  .values[].value","string","是","属性值(文本)","Черный"],
    ["  .values[].dictionary_value_id","int","否","字典值ID(字典类型属性必填)","971082156"],
    ["","","","",""],
    [">>图片(images)","","","",""],
    ["images","array","否","图片URL数组(最多8张)",""],
    ["  .file_name","string","否","文件名(可为空)",""],
    ["  .default","bool","否","是否为主图","true"],
    ["  (或直接传URL字符串)","string","否","系统自动转换为对象格式","https://...jpg"],
    ["","","","",""],
    [">>价格更新字段 (/v4/product/info/prices)","","","",""],
    ["offer_id","string","是","商品ID","SKU-001"],
    ["price","string","是","新售价","561.50"],
    ["old_price","string","否","新原价","699.00"],
    ["min_price","string","否","新最低价","500.00"],
    ["currency_code","string","是","货币代码(固定RUB)","RUB"],
    ["","","","",""],
    [">>库存更新字段 (/v2/product/import/stocks)","","","",""],
    ["offer_id","string","是","商品ID","SKU-001"],
    ["product_id","int","否","OZON商品ID","123456789"],
    ["stock","int","是","库存数量","100"],
    ["warehouse_id","int","否","仓库ID(默认0)","0"],
])

# ═══ Sheet 4: 错误处理 ═══
ws4 = wb.create_sheet("错误处理")
setup(ws4, ["异常类","触发条件","是否重试","处理建议"], [22, 30, 8, 40])
rows(ws4, [
    ["OzonAuthError","HTTP 401/403\n(认证失败/无权限)","否","检查Client-Id和Api-Key是否正确\n确认API权限是否开通"],
    ["OzonRateLimitError","HTTP 429\n(请求过于频繁)","是(2次)","自动指数退避重试\n如持续触发需降低请求频率"],
    ["OzonServerError","HTTP 500/502/503/504\n(服务器错误)","是(2次)","自动重试\n如持续失败需联系OZON技术支持"],
    ["OzonValidationError","HTTP 400/其他4xx\n(参数校验失败)","否","检查请求参数格式\n常见: Items为空/必填字段缺失/类型错误"],
    ["OzonAPIError","网络超时/连接失败\n/其他未知错误","是(2次)","自动重试\n检查网络连接\n确认API地址可访问"],
])

# ═══ Sheet 5: 我们系统的使用方式 ═══
ws5 = wb.create_sheet("系统集成说明")
setup(ws5, ["业务场景","调用的API","调用位置","请求构建方式","注意事项"], [18, 28, 30, 35, 35])
rows(ws5, [
    ["测试API连接","POST /v3/product/list\n(limit=1)","ozon.py account_test()\nozon_api.py test_connectivity()","空filter查1条商品","用于验证Client-Id/Api-Key是否有效"],
    ["同步类目树","POST /v1/description-category/tree","ozon.py api_sync_category_tree()\nozon_api.py get_category_tree()","language=DEFAULT","类目约5000+节点,首次同步较慢"],
    ["同步属性Schema","POST /v1/description-category/attribute","ozon.py api_sync_category_attributes()\nozon_api.py get_category_attributes()","需要同时传category_id和type_id","type_id必填,否则报错"],
    ["同步属性字典值","POST /v1/description-category/attribute/values","ozon.py api_sync_attribute_values()\nozon_api.py get_attribute_values()","limit最大5000","字典类型属性才有字典值"],
    ["发布新商品","POST /v3/product/import","ozon.py listing_publish()\nozon_api.py import_product()","完整商品数据包装为{items:[...]}","必须包含offer_id/name/category_id\n价格和图片建议同时提交"],
    ["查询发布状态","POST /v1/product/import/info","ozon.py (发布后轮询)\nozon_api.py import_product_info()","传入task_id","发布是异步的,需轮询状态"],
    ["更新商品标题/描述","POST /v3/product/import","ozon.py online_product_update()\nozon_api.py import_product()","必须携带完整现有数据\n(name+images+attributes)","不要只发修改字段,否则OZON判定为非正常更新"],
    ["更新商品价格","POST /v4/product/info/prices","ozon.py online_product_update()\nozon_api.py update_product_prices()","{prices:[{offer_id,price,currency_code}]}","专用端点,不影响其他字段"],
    ["更新商品库存","POST /v2/product/import/stocks","ozon.py online_product_update()\nozon_api.py update_product_stocks()","{stocks:[{offer_id,product_id,stock}]}","专用端点,不影响其他字段"],
    ["同步在线商品列表","POST /v3/product/list","ozon.py online_products_sync()\nozon_api.py list_products()","分页遍历,每页100条","last_id分页,非offset分页"],
    ["获取商品详情","POST /v3/product/info/list","ozon.py online_product_sync_detail()\nozon_api.py get_product_info()","传offer_id或product_id","可批量查询多个商品"],
    ["归档/下架商品","POST /v1/product/archive","ozon.py online_product_archive()\nozon_api.py archive_products()","传product_id数组","可批量操作"],
    ["取消归档/上架","POST /v1/product/unarchive","ozon.py online_product_unarchive()\nozon_api.py unarchive_products()","传product_id数组","可批量操作"],
])

path = os.path.join("G:", os.sep, "inventory", "docs", "OZON_Seller_API_Guide.xlsx")
wb.save(path)
print("OK:", path)
