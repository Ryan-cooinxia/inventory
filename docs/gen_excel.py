"""生成 OZON 模块功能清单 Excel"""
import openpyxl, os
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# 样式
hf = Font(bold=True, size=11, color="FFFFFF")
hfill = PatternFill("solid", fgColor="2F5496")
done = PatternFill("solid", fgColor="C6EFCE")
part = PatternFill("solid", fgColor="FFEB9C")
fail = PatternFill("solid", fgColor="FFC7CE")
wrap = Alignment(wrap_text=True, vertical="top")
bdr = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

def write_sheet(ws, headers, widths, rows, status_col=None):
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hf; c.fill = hfill; c.border = bdr
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    for r, rd in enumerate(rows, 2):
        for ci, val in enumerate(rd, 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.alignment = wrap; c.border = bdr
            if status_col and ci == status_col:
                s = str(val)
                if "已完成" in s: c.fill = done
                elif "部分" in s: c.fill = part
                elif "未完成" in s or "废弃" in s: c.fill = fail
            if ci == 1 and str(val).startswith("P"):
                if val == "P0": c.fill = fail
                elif val == "P1": c.fill = part
                else: c.fill = done

# ═══ Sheet 1: 功能清单 ═══
ws1 = wb.active
ws1.title = "OZON模块功能清单"
write_sheet(ws1,
    ["序号","功能模块","页面/功能","路由","完成状态","本次会话改动","存在的问题","优先级"],
    [6, 14, 22, 32, 10, 35, 40, 8],
    [
        [1,"商品采集","采集列表页","/ozon/sources","已完成",
         "新增[导入1688文件夹]和[上传1688 HTML]按钮及面板","无","P0"],
        [2,"商品采集","浏览器插件采集(淘宝/天猫)","/ozon/api/sources/add","已完成",
         "取消H5补采;淘宝/天猫只采标题/SKU/主图;详情图不再报错","无","P0"],
        [3,"商品采集","浏览器插件采集(1688)","/ozon/api/sources/add","部分完成",
         "重写5个专用函数;新增viewData JSON提取;新增DOM调试日志",
         "1688新版页面DOM选择器不稳定;viewData提取待验证","P1"],
        [4,"商品采集","上传1688 HTML文件","/ozon/sources/import-1688-html","已完成",
         "[新增]集成jiyun/1688解析器;上传SingleFile HTML自动解析+下载图片",
         "需安装SingleFile扩展;需克隆jiyun/1688到G:/tools/1688-scraper","P0"],
        [5,"商品采集","导入1688文件夹","/ozon/sources/import-1688-folder","已完成",
         "[新增]读取T_/color_/C_*.jpg + attribute.html自动创建记录",
         "需手动运行jiyun/1688工具再填路径","P2"],
        [6,"商品采集","网页URL采集","/ozon/sources/collect-url","已完成",
         "移除详情图缺失的headless回退","1688 JS渲染内容可能抓取不全","P1"],
        [7,"商品采集","手动粘贴JSON","/ozon/sources/add","已完成","无改动","无","P2"],

        [8,"商品详情","商品详情页","/ozon/sources/<id>","已完成",
         "图片加referrerpolicy防盗链","无","P0"],
        [9,"商品加工","商品加工页(AI生成)","/ozon/processing/<id>","部分完成",
         "修复源图片显示:加缩略图+角色标签+颜色边框",
         "1.AI生成用Mock模拟数据,非真实API\n2.图片无增删功能\n3.AI内容不基于商品数据本土化","P0"],
        [10,"商品适配","商品适配工作台","/ozon/adaptation/<id>","部分完成",
         "图库5组+SKU绑定+人工补图UI+防盗链",
         "1.部分图片加载仍失败\n2.导入的本地图片需通过serve路由","P1"],

        [11,"图片管理","图片过滤/恢复","/ozon/api/source-media/<id>/restore","已完成","无改动","无","P2"],
        [12,"图片管理","人工补图(上传/URL)","/ozon/api/source-media/<id>/upload","已完成",
         "[新增]支持文件上传和URL粘贴","无","P1"],
        [13,"图片管理","图片文件服务","/ozon/uploads/source_media/<id>/<file>","已完成",
         "[新增]Flask send_from_directory","无","P2"],
        [14,"图片管理","图片分类系统","classify_source_image_url()","已完成",
         "清理H5条目;保留main/sku/detail/reject","无","P1"],

        [15,"AI内容","AI加工内容生成","/ozon/processing/<id>/generate","未完成","无改动",
         "1.当前为Mock模拟,未调用真实AI API\n2.俄语标题/卖点/描述需基于商品数据+图片生成\n3.需配置API Key","P0"],
        [16,"AI内容","AI图片识别","/ozon/api/adaptation/<id>/analyze-images","部分完成",
         "无改动","需配置VisionModelConfig","P1"],
        [17,"AI内容","AI智能填充(事实库)","/ozon/api/adaptation/<id>/ai-suggest","部分完成",
         "无改动","需配置API Key","P1"],

        [18,"类目属性","OZON类目选择/推荐","/ozon/api/category/*","已完成","无改动","无","P1"],
        [19,"类目属性","属性映射/间隙分析","/ozon/api/attributes/*","已完成","无改动","无","P2"],

        [20,"刊登发布","刊登草稿管理","/ozon/listings","已完成","无改动","无","P1"],
        [21,"刊登发布","发布到OZON","/ozon/listings/<id>/publish","部分完成",
         "无改动","需配置OZON API账户和Token","P1"],

        [22,"店铺管理","OZON账户管理","/ozon/accounts","已完成","无改动","无","P2"],
        [23,"工具","提示词库","/ozon/prompts","已完成","无改动","无","P2"],
        [24,"工具","定价规则","/ozon/pricing","已完成","无改动","无","P2"],
    ], status_col=5)

# ═══ Sheet 2: 本次改动 ═══
ws2 = wb.create_sheet("本次会话改动汇总")
write_sheet(ws2,
    ["序号","改动文件","类型","改动内容","影响范围"],
    [6, 35, 10, 50, 25],
    [
        [1,"services/ozon_collector.py","删除",
         "删除fetch_taobao_h5_detail_images()约500行;移除h5_main和h5_detail_content规则","H5补采完全移除"],
        [2,"blueprints/ozon.py","重构+新增",
         "删除H5触发逻辑;新增source_selector/reason/linked_sku_name透传;新增4个路由(补图/文件服务/文件夹导入/HTML上传)","采集API+导入功能"],
        [3,"browser-extension/content.js","重构",
         "新增5个1688专用函数;淘宝取消详情采集;Phase3/3.5按平台控制;viewData JSON提取;新版DOM选择器","1688+淘宝采集逻辑"],
        [4,"templates/ozon/adaptation_workspace.html","重构",
         "图库5组+SKU绑定+人工补图面板+referrerpolicy防盗链","适配工作台"],
        [5,"templates/ozon/sources.html","新增",
         "导入1688文件夹面板+上传1688 HTML面板+防盗链","采集列表页"],
        [6,"templates/ozon/source_detail.html","修复","referrerpolicy防盗链","详情页图片"],
        [7,"templates/ozon/processing.html","修复",
         "源图片从占位图标改为缩略图+角色标签+颜色边框","加工页"],
        [8,"G:/tools/1688-scraper/","新增",
         "克隆jiyun/1688工具;安装依赖;创建start_fixed.bat","外部工具集成"],
    ])

# ═══ Sheet 3: 待办 ═══
ws3 = wb.create_sheet("待办事项")
write_sheet(ws3,
    ["优先级","任务","涉及文件","说明"],
    [8, 30, 30, 50],
    [
        ["P0","AI内容生成:Mock转真实API调用",
         "ozon.py processing_generate()",
         "当前为模拟数据;需调用OpenAI/DeepSeek生成真实俄语标题/卖点/描述;需配置API Key"],
        ["P0","AI Prompt优化:基于商品数据本土化",
         "ozon.py + AI prompts",
         "AI应基于:商品属性+SKU信息+图片识别+采集内容 生成本土化俄语描述"],
        ["P0","加工页图片增删功能",
         "processing.html + ozon.py",
         "支持:删除不需要的图片/添加新图片/调整图片角色(主图/SKU/详情)"],
        ["P1","1688插件采集稳定性",
         "content.js",
         "viewData JSON提取待验证;DOM选择器需适配更多页面版本"],
        ["P1","适配页图片加载优化",
         "adaptation_workspace.html",
         "考虑后端代理下载图片到本地,避免前端referrer问题"],
        ["P2","图片质量自动检测",
         "ozon_collector.py",
         "自动检测服务图标/资质认证/营业执照等非商品图片"],
        ["P2","多1688页面版本兼容",
         "content.js",
         "测试不同类型1688商品页确保选择器覆盖率"],
    ])

os.makedirs("G:/inventory/docs", exist_ok=True)
path = "G:/inventory/docs/OZON模块功能清单_20260620.xlsx"
wb.save(path)
print("OK:", path)
