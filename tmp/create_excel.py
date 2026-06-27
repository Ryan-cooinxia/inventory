import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

header_font = Font(name='Microsoft YaHei', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
done_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
todo_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
normal_font = Font(name='Microsoft YaHei', size=10)
bold_font = Font(name='Microsoft YaHei', size=10, bold=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
wrap_align = Alignment(wrap_text=True, vertical='top')

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font; cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_row(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = normal_font; cell.border = thin_border; cell.alignment = wrap_align

# ====== Sheet 1: 项目总览 ======
ws = wb.active
ws.title = '项目总览'
ws.column_dimensions['A'].width = 18; ws.column_dimensions['B'].width = 60
for i,h in enumerate(['项目','内容'],1): ws.cell(row=1,column=i,value=h)
style_header(ws,1,2)
data = [
    ['项目名称','产品母图自动识别与目标级抠图系统'],
    ['所属系统','仓库记账系统 - OZON运营模块'],
    ['当前阶段','P0 基础闭环已完成，P1 待开发'],
    ['最后更新','2026-06-25'],
    ['核心原则','AI只负责定位商品，正式母图像素来自原图，禁止AI重绘；视觉模型(语义)与分割模型(像素)职责分离；人工确认后才可设为正式产品母图；所有数据按current_user隔离'],
    ['工作流','原图 → 视觉识别主商品/排除区 → 搜索框自动扩大 → rembg only_mask → 正向点选连通区域 → mask反推真实bbox → 原图RGB+mask合成 → 质量门禁 → 人工批准 → 透明母图'],
]
for i,(k,v) in enumerate(data,2):
    ws.cell(row=i,column=1,value=k).font=bold_font
    ws.cell(row=i,column=2,value=v).font=normal_font
    style_row(ws,i,2)

# ====== Sheet 2: 阶段规划 ======
ws = wb.create_sheet('阶段规划')
ws.column_dimensions['A'].width=8; ws.column_dimensions['B'].width=16; ws.column_dimensions['C'].width=12; ws.column_dimensions['D'].width=65
for i,h in enumerate(['阶段','名称','状态','内容'],1): ws.cell(row=1,column=i,value=h)
style_header(ws,1,4)
for i,(p,n,s,c) in enumerate([
    ('P0','基础闭环','已完成','自动识别+搜索框分割+rembg only_mask+连通区域选择+排除框保护+质量门禁+人工批准'),
    ('P1','分割升级','待开发','SAM 2 box prompt分割+正/负向点+蒙版画笔修正(擦除/恢复)+边缘背景色去污染'),
    ('P2','批量与联动','待开发','多配件实例分割+SKU级母图+批量母图+Composite生图(AI只生成背景,后端合成透明母图)'),
],2):
    for j,v in enumerate([p,n,s,c],1): ws.cell(row=i,column=j,value=v)
    style_row(ws,i,4)

# ====== Sheet 3: P0完成清单 ======
ws = wb.create_sheet('P0完成清单')
ws.column_dimensions['A'].width=12; ws.column_dimensions['B'].width=40; ws.column_dimensions['C'].width=55
for i,h in enumerate(['分类','项目','说明'],1): ws.cell(row=1,column=i,value=h)
style_header(ws,1,3)
items = [
    ('数据库','OzonProductCutout(26字段)','透明PNG/mask/质量评分/目标框/分割方式/像素验证/修订版/边缘评分'),
    ('数据库','OzonProductSubjectDetection(15字段)','视觉检测历史:主商品/排除区/置信度/正向点/原始响应'),
    ('数据库','OzonImagePlan(14字段)','图片方案:listing/aplus/目标市场/语言'),
    ('数据库','OzonImageReference(8字段)','参考图关联:role/priority/required'),
    ('','',''),
    ('服务','product_cutout.py V3.2','搜索框分割+rembg only_mask+连通区域选择+mask反推bbox+像素验证+质量检查+文字检测+棋盘格预览'),
    ('服务','product_subject_detector.py','复用VisionModelConfig(千问VL/OpenAI):返回主商品/排除区/正向点/置信度'),
    ('服务','ecommerce_image_reference.py','按slot role选参考图:SKU优先/package不冒充main'),
    ('服务','image_generation.py','多Provider(Seedream/GPT/Wanxiang):参考图+请求快照+Seedream独立适配器'),
    ('服务','ecommerce_image_skill.py','Prompt构建:marketplace/language参数,OZON默认俄语'),
    ('','',''),
    ('路由','GET /product-cutout/<source_id>','产品母图准备页面'),
    ('路由','POST /product-cutout/<media_id>/create','执行抠图(支持targets JSON+自动识别结果)'),
    ('路由','POST /product-cutout/<media_id>/detect-subject','视觉模型自动识别商品主体'),
    ('路由','POST /product-cutout/<cutout_id>/approve','确认母图(含强制质量门禁:无目标/质量不过/像素修改/框外残留/rembg_full→拒绝)'),
    ('路由','POST /product-cutout/<cutout_id>/reject','拒绝母图'),
    ('路由','GET /uploads/cutouts/<path>','访问抠图文件(透明PNG/mask/预览)'),
    ('路由','GET /uploads/ai_generated/<path>','AI生图文件访问'),
    ('路由','GET /sources/<id>/download-images','一键下载全图ZIP'),
    ('路由','POST /api/models/test-image-gen','测试图片生成配置(脱敏请求体)'),
    ('','',''),
    ('前端','product_cutout.html','自动识别按钮+目标框拖拽缩放+棋盘格/白/灰/黑背景切换+母图确认/拒绝+下载透明PNG'),
    ('前端','sources.html','采集列表每行「✂️ 母图」入口+「📥」一键下载ZIP'),
    ('','',''),
    ('核心技术','rembg only_mask','只出mask不重绘,原图RGB+mask合成,像素差异验证=0'),
    ('核心技术','搜索框自动扩大','视觉bbox→搜索框(左30%/上15%/右12%/下8%),覆盖完整产品'),
    ('核心技术','正向点连通区域选择','搜索框内rembg→连通区域标记→正向点匹配→只保留商品区域'),
    ('核心技术','排除框保护商品','排除框只清零非商品连通区域,不切割商品mask'),
    ('核心技术','mask触边自动扩框重试','检测mask触碰搜索框边缘→扩大重试→最多2次'),
    ('核心技术','坐标缩放','页面显示坐标×scale→原图坐标'),
    ('核心技术','质量计算修复','binary.mean()不再/255,rembg_full最高40分且不pass'),
    ('核心技术','批准强制门禁','无目标/质量不过/像素修改/框外残留/rembg_full→拒绝批准'),
    ('核心技术','文字/Logo检测','边缘密度+局部对比度检测广告文字,扣分+警告'),
    ('核心技术','Seedream适配器','image数组参数+watermark=false+size=2K+response_format=url'),
]
for i,(a,b,c) in enumerate(items,2):
    ws.cell(row=i,column=1,value=a); ws.cell(row=i,column=2,value=b); ws.cell(row=i,column=3,value=c)
    style_row(ws,i,3)

# ====== Sheet 4: 当前问题 ======
ws = wb.create_sheet('当前问题')
ws.column_dimensions['A'].width=6; ws.column_dimensions['B'].width=14; ws.column_dimensions['C'].width=50; ws.column_dimensions['D'].width=8; ws.column_dimensions['E'].width=14; ws.column_dimensions['F'].width=10; ws.column_dimensions['G'].width=45
for i,h in enumerate(['#','分类','问题','严重度','状态','归属','备注'],1): ws.cell(row=1,column=i,value=h)
style_header(ws,1,7)
issues = [
    (1,'视觉识别','视觉bbox[446,102,800,800]比产品实际(~x=355)窄90px','中','V3.2已修复','P0','搜索框左扩30%覆盖完整产品,实际分割框[358,83,784,753]'),
    (2,'排除框','排除框x延伸到385-437,与商品左边缘重叠','中','V3.2已修复','P0','排除框只清零非商品连通区域'),
    (3,'分割质量','rembg边缘复杂时残留/缺失(白色商品/透明材质/毛发/细线缆)','低','待P1','P1','期待SAM 2改善'),
    (4,'识别延迟','视觉模型调用3-5秒','低','可接受','-','用户等待在合理范围'),
    (5,'白色商品','白色商品在白底时边缘被rembg吃掉','高','待P1','P1','需SAM 2或边缘增强'),
    (6,'玻璃/透明','玻璃/透明/反光材质分割困难','中','待P2','P2','需特殊处理'),
    (7,'细结构','天线/线缆/毛发在mask清理时被误删','中','待P1','P1','需改进蒙版清理逻辑'),
    (8,'多商品','多商品堆叠时无法区分主商品','中','待P1','P1','需多实例分割'),
    (9,'SAM 2','未接入SAM 2分割','高','待P1','P1','P1核心:替代rembg'),
    (10,'蒙版修正','无蒙版画笔修正(擦除/恢复)','高','待P1','P1','用户需手工修正mask'),
    (11,'边缘去色','无边缘背景色去污染','中','待P1','P1','红/粉色背景边缘残留'),
    (12,'SKU母图','不同SKU无法分别管理母图','高','待P2','P2','多颜色/款式需独立母图'),
    (13,'批量','无法一键对多图抠图','中','待P2','P2','批量母图准备'),
    (14,'Composite','AI生图Composite模式未实现','高','待P2','P2','AI生成背景+后端合成母图'),
    (15,'联动','母图未与Seedream生图联动','低','待P2','P2','母图确认后自动用于生图'),
]
for i,(a,b,c,d,e,f,g) in enumerate(issues,2):
    for j,v in enumerate([a,b,c,d,e,f,g],1): ws.cell(row=i,column=j,value=v)
    style_row(ws,i,7)

# ====== Sheet 5: 技术架构 ======
ws = wb.create_sheet('技术架构')
ws.column_dimensions['A'].width=12; ws.column_dimensions['B'].width=45; ws.column_dimensions['C'].width=55
for i,h in enumerate(['层级','文件/路由','功能描述'],1): ws.cell(row=1,column=i,value=h)
style_header(ws,1,3)
arch = [
    ('数据模型','models.py','OzonProductCutout(26字段)+OzonProductSubjectDetection(15字段)+OzonImagePlan(14字段)+OzonImageReference(8字段)+OzonImageSlot(27字段)+OzonImageCandidate(32字段)'),
    ('迁移','app.py migrate_ozon_image_schema()','幂等ALTER TABLE,PRAGMA检测缺失字段,逐列添加'),
    ('服务','services/product_cutout.py(~450行)','搜索框分割+rembg only_mask+连通区域选择+mask反推bbox+像素验证+质量检查+文字检测+棋盘格预览'),
    ('服务','services/product_subject_detector.py(~220行)','复用VisionModelConfig,千问VL/OpenAI兼容,返回主商品/排除区/正向点'),
    ('服务','services/ecommerce_image_reference.py(~140行)','按slot role选参考图,SKU优先'),
    ('服务','services/image_generation.py(~650行)','多Provider生成,Seedream适配器,GPT Image,请求快照'),
    ('服务','services/ecommerce_image_skill.py(~200行)','Prompt构建,OZON默认俄语,marketplace参数'),
    ('存储','uploads/source_media/<id>/','采集原图'),
    ('存储','uploads/cutouts/<media_id>/','母图:cutout_PNG+mask_raw+mask_clean+preview_JPG'),
    ('存储','uploads/ai_generated/draft_<id>/slot_<n>/','AI生图候选'),
]
for i,(a,b,c) in enumerate(arch,2):
    ws.cell(row=i,column=1,value=a); ws.cell(row=i,column=2,value=b); ws.cell(row=i,column=3,value=c)
    style_row(ws,i,3)

# ====== Sheet 6: 开发日志 ======
ws = wb.create_sheet('开发日志')
ws.column_dimensions['A'].width=14; ws.column_dimensions['B'].width=10; ws.column_dimensions['C'].width=8; ws.column_dimensions['D'].width=65
for i,h in enumerate(['日期','版本','类型','内容'],1): ws.cell(row=1,column=i,value=h)
style_header(ws,1,4)
log = [
    ('2026-06-24','V1','feat','产品母图基础模型+rembg整图抠图'),
    ('2026-06-24','V2','feat','目标框+rembg_crop+框外强制透明+文字/Logo检测'),
    ('2026-06-24','V3','feat','rembg only_mask只出mask+原图像素零修改+像素真实性验证'),
    ('2026-06-25','V3.1','fix','坐标缩放修复+禁止复杂图rembg_full+质量计算修复+排除框生效+批准门禁'),
    ('2026-06-25','V3.2','fix','视觉bbox降级为种子框+搜索框自动扩大(左30%)+正向点选连通区域+排除框保护商品+mask触边重试'),
    ('2026-06-25','P0-1','feat','视觉模型自动识别:OzonProductSubjectDetection+detect-subject路由+🔍自动识别按钮'),
    ('2026-06-25','-','doc','需求文档(Excel)创建'),
]
for i,(a,b,c,d) in enumerate(log,2):
    ws.cell(row=i,column=1,value=a); ws.cell(row=i,column=2,value=b); ws.cell(row=i,column=3,value=c); ws.cell(row=i,column=4,value=d)
    style_row(ws,i,4)

output = r'G:\inventory\designs\product-cutout-flow\design_handoff\需求文档.xlsx'
wb.save(output)
print(f'Done: {output}')
