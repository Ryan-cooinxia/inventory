# blueprints/ai_import.py
from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash
from flask_login import login_required, current_user
from models import Product, UserApiKey
from crypto_utils import encrypt_api_key, decrypt_api_key
import pdfplumber
import openai
import io
import json
import re
import openpyxl

ai_bp = Blueprint('ai_import', __name__)

# ---------- AI 设置页面 ----------
@ai_bp.route('/ai-settings', methods=['GET', 'POST'])
@login_required
def ai_settings():
    if request.method == 'POST':
        api_key = request.form.get('api_key', '').strip()
        provider = request.form.get('provider', 'deepseek')
        if not api_key:
            flash('API Key 不能为空', 'danger')
            return redirect(url_for('ai_import.ai_settings'))
        encrypted = encrypt_api_key(api_key)
        record, created = UserApiKey.get_or_create(
            user=current_user,
            defaults={'api_key': encrypted, 'api_provider': provider}
        )
        if not created:
            record.api_key = encrypted
            record.api_provider = provider
            record.save()
        flash('API Key 保存成功', 'success')
        return redirect(url_for('ai_import.ai_settings'))

    key_info = None
    record = UserApiKey.get_or_none(UserApiKey.user == current_user)
    if record:
        decrypted = decrypt_api_key(record.api_key)
        if len(decrypted) > 8:
            masked = decrypted[:4] + '****' + decrypted[-4:]
        else:
            masked = '****'
        key_info = {'provider': record.api_provider, 'masked_key': masked}
    return render_template('ai_settings.html', key_info=key_info)

@ai_bp.route('/ai-settings/delete', methods=['POST'])
@login_required
def ai_settings_delete():
    UserApiKey.delete().where(UserApiKey.user == current_user).execute()
    flash('API Key 已删除', 'info')
    return redirect(url_for('ai_import.ai_settings'))

# ---------- AI 导入页面 ----------
@ai_bp.route('/ai-import')
@login_required
def ai_import_page():
    return render_template('ai_import.html')

# ---------- 上传并解析（核心功能） ----------
@ai_bp.route('/api/ai-import/upload', methods=['POST'])
@login_required
def ai_import_upload():
    key_record = UserApiKey.get_or_none(UserApiKey.user == current_user)
    if not key_record:
        return jsonify({'error': '请先在系统设置中配置 AI API Key'}), 400

    api_key = decrypt_api_key(key_record.api_key)
    provider = key_record.api_provider

    file = request.files.get('file')
    if not file:
        return jsonify({'error': '未上传文件'}), 400

    # ---------- 提取文本 ----------
    try:
        if file.filename.endswith('.pdf'):
            raw_text = extract_pdf_text(file)
        elif file.filename.endswith(('.xlsx', '.xls')):
            raw_text = extract_excel_text(file)
        else:
            return jsonify({'error': '仅支持 PDF 和 Excel 文件'}), 400
    except Exception as e:
        return jsonify({'error': f'文件解析失败：{str(e)}'}), 400

    if not raw_text.strip():
        return jsonify({'error': '未从文件中提取到任何文本'}), 400

    # 初始化 OpenAI 客户端
    try:
        client = openai.OpenAI(
            api_key=api_key,
            base_url="https://api.deepseek.com" if provider == 'deepseek' else None
        )
    except Exception as e:
        return jsonify({'error': f'AI 初始化失败：{str(e)}'}), 500

    # ---------- 单次调用 AI ----------
    def call_ai(chunk_text, is_last_chunk=False):
        prompt = f"""
你是一个智能产品信息提取助手。请从以下内容中提取**所有可能的产品**，输出一个 JSON 对象：{{"products": [...]}}。
products 数组中每个元素必须包含：
- name (产品名称/型号，**必须**，如果没有明确名称，则从描述中归纳，不可省略)
- spec (规格/包装，保留颜色、包装类型等)
- description (产品说明，无则为 "")
- unit (单位，默认"盒")
- unit_price (单价数字，无则 null)
- quantity (数量数字，无则 null)
- category1 (一级分类，推断)
- category2 (二级分类，推断)
- brand (品牌，推断，无则 null)

要求：
- 只输出 JSON 对象。
- 必须提取**所有**能找到的产品，不要遗漏任何一行产品信息。
- 某个字段无法确定则设为 null 或空字符串。

文本{'(最后一段)' if is_last_chunk else ''}：
{chunk_text[:8000]}
"""
        try:
            resp = client.chat.completions.create(
                model="deepseek-chat",
                messages=[{"role": "user", "content": prompt}],
                temperature=0.1,
                max_tokens=4096,
                response_format={"type": "json_object"}
            )
            result_text = resp.choices[0].message.content.strip()
            if result_text.startswith('```'):
                result_text = result_text.split('\n', 1)[1].rsplit('\n', 1)[0]
            return result_text
        except Exception as e:
            print(f"AI 调用错误: {e}")
            return None

    # ---------- 按长度拆分文本 ----------
    MAX_CHUNK_SIZE = 5000
    if len(raw_text) <= MAX_CHUNK_SIZE:
        chunks = [raw_text]
    else:
        lines = raw_text.split('\n')
        chunks = []
        current = ''
        for line in lines:
            if len(current) + len(line) > MAX_CHUNK_SIZE and current:
                chunks.append(current)
                current = line + '\n'
            else:
                current += line + '\n'
        if current:
            chunks.append(current)

    # ---------- 解析所有分块 ----------
    all_products = []
    for idx, chunk in enumerate(chunks[:8]):  # 最多处理8段，避免超时
        result_text = call_ai(chunk, is_last_chunk=(idx == len(chunks) - 1))
        if result_text is None:
            continue
        products = _parse_ai_json(result_text)
        if products:
            all_products.extend(products)

    if not all_products:
        preview = raw_text[:300].replace('\n', ' ')
        return jsonify({'error': f'AI 未识别到任何产品。文件开头预览：{preview}...'}), 400

    return jsonify({'products': all_products})


# ---------- 确认导入 ----------
@ai_bp.route('/api/ai-import/confirm', methods=['POST'])
@login_required
def ai_import_confirm():
    data = request.get_json()
    products = data.get('products', [])
    if not products:
        return jsonify({'error': '没有产品数据'}), 400
    count = 0
    for item in products:
        try:
            Product.create(
                name=item['name'],
                spec=item.get('spec') or None,
                description=item.get('description') or None,
                unit=item.get('unit') or '盒',
                brand=item.get('brand') or None,
                category1=item.get('category1') or None,
                category2=item.get('category2') or None,
                user=current_user
            )
            count += 1
        except Exception:
            continue
    return jsonify({'count': count})


# ---------- 辅助函数：提取 PDF / Excel 文本 ----------
def extract_pdf_text(file):
    with pdfplumber.open(io.BytesIO(file.read())) as pdf:
        text = ''
        for page in pdf.pages:
            text += page.extract_text() or ''
    return text

def extract_excel_text(file):
    wb = openpyxl.load_workbook(io.BytesIO(file.read()))
    text = ''
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            row_text = ' '.join([str(cell) for cell in row if cell is not None])
            text += row_text + '\n'
    return text


# ---------- JSON 解析与修复 ----------
def _parse_ai_json(text):
    """尝试从可能不完美的 AI 返回中提取产品列表"""
    # 1. 直接解析
    try:
        data = json.loads(text)
        return _extract_products(data)
    except Exception:
        pass

    # 2. 修复常见问题后解析
    repaired = text.strip()
    # 补全缺失的尾部大括号
    if repaired.count('{') > repaired.count('}'):
        repaired += '}' * (repaired.count('{') - repaired.count('}'))
    # 移除对象间多余的逗号
    repaired = re.sub(r',\s*}', '}', repaired)
    repaired = re.sub(r',\s*]', ']', repaired)
    try:
        data = json.loads(repaired)
        return _extract_products(data)
    except Exception:
        pass

    # 3. 逐行查找 JSON 对象（正则匹配）
    products = []
    # 匹配嵌套对象（简化版）
    pattern = r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}'
    matches = re.findall(pattern, text)
    for match in matches:
        try:
            obj = json.loads(match)
            if isinstance(obj, dict) and 'name' in obj and obj['name']:
                products.append(obj)
        except Exception:
            continue
    if products:
        return products

    # 4. 尝试提取 products 数组的内容
    try:
        arr_match = re.search(r'"products"\s*:\s*(\[.*?\])', text, re.DOTALL)
        if arr_match:
            arr_text = arr_match.group(1)
            arr_text_repaired = re.sub(r',\s*]', ']', arr_text)
            products = json.loads(arr_text_repaired)
            if isinstance(products, list):
                return products
    except Exception:
        pass

    return []


def _extract_products(data):
    """从解析出的对象中提取产品列表"""
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        if 'products' in data:
            return data['products']
        # 取第一个列表值
        for val in data.values():
            if isinstance(val, list):
                return val
    return []