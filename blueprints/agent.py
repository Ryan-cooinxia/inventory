# blueprints/agent.py
from flask import Blueprint, request, jsonify, render_template
from flask_login import login_required, current_user
from models import (
    UserApiKey, Product, Customer, Supplier,
    CustomerOrder, SalesOrder, PurchaseOrder, SupplierOrder,
    PurchaseOrderItem, SalesOrderItem
)
from crypto_utils import decrypt_api_key
from peewee import fn
import openai
import json
import datetime
import base64
import io
import openpyxl
import pdfplumber

agent_bp = Blueprint('agent', __name__)

# ===================== 系统内置工具函数 =====================

def get_all_inventory(user):
    """返回所有库存不为0的产品列表"""
    products = Product.select().where(Product.user == user)
    result = []
    for p in products:
        total_in = (PurchaseOrderItem.select(fn.SUM(PurchaseOrderItem.quantity))
                    .join(PurchaseOrder).where((PurchaseOrderItem.product == p) & (PurchaseOrder.user == user))
                    .scalar()) or 0
        total_out = (SalesOrderItem.select(fn.SUM(SalesOrderItem.quantity))
                     .join(SalesOrder).where((SalesOrderItem.product == p) & (SalesOrder.user == user))
                     .scalar()) or 0
        stock = total_in - total_out
        if stock != 0:
            result.append(f"{p.sku or ''} {p.name}：库存{stock}")
    return "\n".join(result) if result else "当前没有任何产品库存。"

def search_inventory(keyword, user):
    """模糊搜索产品并返回库存"""
    products = Product.select().where(
        (Product.name.contains(keyword) | Product.sku.contains(keyword)) & (Product.user == user)
    )
    if not products.exists():
        return f"未找到包含'{keyword}'的产品。"
    result = []
    for p in products:
        total_in = (PurchaseOrderItem.select(fn.SUM(PurchaseOrderItem.quantity))
                    .join(PurchaseOrder).where((PurchaseOrderItem.product == p) & (PurchaseOrder.user == user))
                    .scalar()) or 0
        total_out = (SalesOrderItem.select(fn.SUM(SalesOrderItem.quantity))
                     .join(SalesOrder).where((SalesOrderItem.product == p) & (SalesOrder.user == user))
                     .scalar()) or 0
        stock = total_in - total_out
        result.append(f"{p.name}（{p.sku or ''}）：库存{stock}")
    return "\n".join(result)

def get_unfinished_orders(user):
    """获取未完成的客户订单"""
    orders = CustomerOrder.select().where(
        (CustomerOrder.user == user) & (CustomerOrder.status != 'shipped') & (CustomerOrder.status != 'completed')
    )
    if not orders.exists():
        return "目前没有未完成的客户订单。"
    result = []
    for o in orders:
        shipped = (SalesOrder.select(fn.SUM(SalesOrder.total_amount))
                   .where((SalesOrder.customer_order == o) & (SalesOrder.user == user)).scalar()) or 0
        remaining = o.total_amount - shipped
        result.append(f"订单#{o.id} {o.customer.name} 总¥{o.total_amount:.2f} 已发¥{shipped:.2f} 未发¥{remaining:.2f}")
    return "\n".join(result)

def get_unreceived_supplier_orders(user):
    """获取未完成的供应商订单"""
    orders = SupplierOrder.select().where(
        (SupplierOrder.user == user) & (SupplierOrder.status != 'received') & (SupplierOrder.status != 'completed')
    )
    if not orders.exists():
        return "目前没有未完成的供应商订单。"
    result = []
    for o in orders:
        received = (PurchaseOrder.select(fn.SUM(PurchaseOrder.total_amount))
                    .where((PurchaseOrder.supplier_order == o) & (PurchaseOrder.user == user)).scalar()) or 0
        remaining = o.total_amount - received
        result.append(f"订单#{o.id} {o.supplier.name} 总¥{o.total_amount:.2f} 已收¥{received:.2f} 未收¥{remaining:.2f}")
    return "\n".join(result)

def get_top_products(user, days=30):
    """最近N天销量前5的产品"""
    since_date = datetime.date.today() - datetime.timedelta(days=days)
    query = (SalesOrderItem
             .select(SalesOrderItem.product, fn.SUM(SalesOrderItem.quantity).alias('total_qty'))
             .join(SalesOrder)
             .where((SalesOrder.order_date >= since_date) & (SalesOrder.user == user))
             .group_by(SalesOrderItem.product)
             .order_by(fn.SUM(SalesOrderItem.quantity).desc())
             .limit(5))
    result = []
    for row in query:
        product = Product.get_or_none(Product.id == row.product_id)
        if product:
            result.append(f"{product.name}：销量{row.total_qty}")
    return "\n".join(result) if result else f"最近{days}天没有销售记录。"

def get_recent_arrival_ranking(user, days=30):
    """最近N天内到货数量最多的产品排行"""
    since_date = datetime.date.today() - datetime.timedelta(days=days)
    query = (PurchaseOrderItem
             .select(PurchaseOrderItem.product,
                     fn.SUM(PurchaseOrderItem.quantity).alias('total_qty'))
             .join(PurchaseOrder)
             .where((PurchaseOrder.order_date >= since_date) & (PurchaseOrder.user == user))
             .group_by(PurchaseOrderItem.product)
             .order_by(fn.SUM(PurchaseOrderItem.quantity).desc())
             .limit(10))
    result = []
    for row in query:
        product = Product.get_or_none(Product.id == row.product_id)
        if product:
            result.append(f"{product.name}：到货 {row.total_qty} 个")
    return "\n".join(result) if result else f"最近{days}天内没有到货记录。"

# 工具映射
TOOL_MAP = {
    "get_all_inventory": get_all_inventory,
    "search_inventory": search_inventory,
    "get_unfinished_orders": get_unfinished_orders,
    "get_unreceived_supplier_orders": get_unreceived_supplier_orders,
    "get_top_products": get_top_products,
    "get_recent_arrival_ranking": get_recent_arrival_ranking,
}

# ===================== 文件解析函数 =====================

def extract_excel_text(file_bytes):
    """从 Excel 文件中提取文本"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    text = ''
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            row_text = ' '.join([str(cell) for cell in row if cell is not None])
            text += row_text + '\n'
    return text

def extract_pdf_text(file_bytes):
    """从 PDF 文件中提取文本"""
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        text = ''
        for page in pdf.pages:
            text += page.extract_text() or ''
    return text

# ===================== 系统提示词 =====================

SYSTEM_PROMPT = """
你是一个专业的仓库管理AI助手。你可以直接调用系统数据来回答用户问题，也可以分析用户上传的Excel/PDF文件。

## 系统内置工具（调用时输出JSON数组）
1. get_all_inventory() - 获取所有有库存的产品列表
2. search_inventory(keyword) - 根据关键词搜索产品库存
3. get_unfinished_orders() - 获取未完成的客户订单
4. get_unreceived_supplier_orders() - 获取未完成的供应商订单
5. get_top_products(days) - 获取最近指定天数的畅销产品
6. get_recent_arrival_ranking(days) - 获取最近指定天数到货数量最多的产品排行

调用格式示例：[{"tool": "get_all_inventory", "args": {}}]

## 文件分析规则（当用户上传文件并要求分析时）
如果用户上传了文件并要求分析到货速度，请从文件内容中提取数据，并按照以下标准分析：
- 每行数据应包含：产品名称、下单日期、到货日期（或入库日期）、数量
- 计算每个产品的平均到货天数 = (到货日期 - 下单日期) 的总和 / 订单数
- 到货速度评级：平均到货天数 ≤ 5天 为“快”，5-15天 为“中等”，>15天 为“慢”
- 输出格式：先给出总体结论，然后列出每个产品的平均天数、评级、涉及订单数，最后给出建议

如果用户没有上传文件，则根据提问调用内置工具。

## 通用规则
- 如果用户的问题需要查询数据，你必须输出一个 JSON 数组，即使只有一个工具也要用数组包裹。
- 如果用户只是闲聊或咨询能力范围外的事，直接回复自然语言。
- 当收到工具返回的数据后，你要用自然语言进行解释，让用户容易理解，并主动提供分析建议。
- 如果用户问“库存”、“还有什么货”等，调用 get_all_inventory()。
- 如果用户问具体产品库存，调用 search_inventory(keyword="产品名")。
- 如果用户问“哪些订单还没完成”，调用 get_unfinished_orders()。
- 如果用户问“最近什么好卖”，调用 get_top_products(days=30)。
- 如果用户问“哪些产品到货快”、“到货速度”、“最近什么到货多”，调用 get_recent_arrival_ranking(days=30)。
- 如果用户的问题需要综合多个数据源，你可以一次返回多个工具调用，我会并行执行它们，然后你综合分析结果。
- 对于文件分析，不需要调用工具，直接基于文件内容回答。
"""

# ===================== 路由 =====================

@agent_bp.route('/api/agent/chat', methods=['POST'])
@login_required
def agent_chat():
    key_record = UserApiKey.get_or_none(UserApiKey.user == current_user)
    if not key_record:
        return jsonify({'error': '请先在AI设置中配置API Key'}), 400

    api_key = decrypt_api_key(key_record.api_key)
    provider = key_record.api_provider

    # 检查是否为文件上传（multipart）
    if request.content_type and 'multipart/form-data' in request.content_type:
        message = request.form.get('message', '').strip()
        file = request.files.get('file')
        if not file:
            return jsonify({'error': '未收到文件'}), 400
        # 解析文件
        try:
            file_bytes = file.read()
            if file.filename.endswith(('.xlsx', '.xls')):
                file_text = extract_excel_text(file_bytes)
            elif file.filename.endswith('.pdf'):
                file_text = extract_pdf_text(file_bytes)
            else:
                return jsonify({'error': '仅支持 Excel 和 PDF 文件'}), 400
        except Exception as e:
            return jsonify({'error': f'文件解析失败：{str(e)}'}), 400

        # 构建消息
        user_message = f"{message}\n\n以下是从上传文件 {file.filename} 中提取的数据：\n{file_text[:3000]}"
        return _process_agent_reply(user_message, None, api_key, provider)

    # JSON 请求（文本/图片）
    data = request.get_json()
    message = data.get('message', '').strip()
    image_base64 = data.get('image', None)

    if not message and not image_base64:
        return jsonify({'error': '请输入消息或上传图片'}), 400

    return _process_agent_reply(message, image_base64, api_key, provider)


def _process_agent_reply(message, image_base64, api_key, provider):
    """处理对话并返回回复"""
    messages = [{"role": "system", "content": SYSTEM_PROMPT}]
    user_content = []
    if message:
        user_content.append({"type": "text", "text": message})
    if image_base64:
        user_content.append({"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{image_base64}"}})
    messages.append({"role": "user", "content": user_content})

    try:
        client = openai.OpenAI(
            api_key=api_key,
            base_url="https://api.deepseek.com" if provider == 'deepseek' else None
        )

        # 第一步：获取AI的原始回复
        response = client.chat.completions.create(
            model="deepseek-chat",
            messages=messages,
            temperature=0.2,
            max_tokens=1024
        )
        ai_text = response.choices[0].message.content.strip()

        # 尝试解析工具调用
        tool_calls = []
        try:
            cleaned = ai_text.strip()
            if cleaned.startswith('```'):
                cleaned = cleaned.split('\n', 1)[1].rsplit('\n', 1)[0]
            # 移除可能的 BOM 或不可见字符
            cleaned = cleaned.strip().lstrip('\ufeff')
            parsed = json.loads(cleaned)
            if isinstance(parsed, list):
                tool_calls = parsed
            elif isinstance(parsed, dict) and 'tool' in parsed:
                tool_calls = [parsed]
        except Exception as e:
            print(f"工具调用解析失败: {e}")   # 可在服务器日志中查看具体错误
            tool_calls = []

        # 如果有工具调用
        if tool_calls:
            tool_results = []
            for call in tool_calls:
                tool_name = call.get('tool')
                args = call.get('args', {})
                if tool_name in TOOL_MAP:
                    try:
                        result_text = TOOL_MAP[tool_name](current_user, **args)
                        tool_results.append(f"工具 {tool_name} 结果：\n{result_text}")
                    except Exception as e:
                        tool_results.append(f"工具 {tool_name} 执行失败：{str(e)}")
                else:
                    tool_results.append(f"未知工具：{tool_name}")

            messages.append({"role": "assistant", "content": ai_text})
            combined_results = "\n\n".join(tool_results)
            messages.append({"role": "user", "content": f"以下是数据查询结果，请用自然语言向用户解释，并主动给出分析建议：\n{combined_results}"})

            response2 = client.chat.completions.create(
                model="deepseek-chat",
                messages=messages,
                temperature=0.6,
                max_tokens=2048
            )
            final_text = response2.choices[0].message.content.strip()
            return jsonify({'reply': final_text})

        # 非工具调用
        return jsonify({'reply': ai_text})

    except Exception as e:
        return jsonify({'error': f'AI调用失败：{str(e)}'}), 500
    
@agent_bp.route('/ai-assistant')
@login_required
def ai_assistant_page():
    """独立AI助手页面"""
    return render_template('ai_assistant.html')