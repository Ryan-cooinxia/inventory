"""OZON 官方 Excel 模板服务层

提供模板解析、草稿字段映射、Excel 生成功能。
"""
import os
import hashlib
import json
import datetime
import io

import openpyxl
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════
# 模板解析
# ═══════════════════════════════════════════════════════════════

def inspect_template(file_path, original_filename):
    """
    解析上传的 OZON 官方 Excel 模板，提取表头结构、必填列、验证规则。

    返回:
        {
            "schema_hash": "a1b2c3d4",
            "headers": [{"index": 0, "letter": "A", "header": "货号", "is_required": true}, ...],
            "required_columns": [{"index": 0, "letter": "A", "header": "货号"}, ...],
            "sheet_names": ["模板", "示例", ...],
            "data_start_row": 5,
            "header_row": 2,
            "validations": [...]
        }

    Raises:
        ValueError: 模板缺少必需的 "模板" sheet
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)

    sheet_names = wb.sheetnames
    if '模板' not in sheet_names:
        wb.close()
        raise ValueError(f"上传的文件缺少「模板」工作表。当前 Sheet: {', '.join(sheet_names)}。请确认这是 OZON 官方模板。")

    ws = wb['模板']

    # OZON 官方模板固定：第 2 行表头，第 5 行数据起始
    # 第 3-4 行是说明/示例行，不写入商品数据
    header_row = 2
    data_start_row = 5

    # 尝试从第 5 行「类型*」列读取示例值，辅助类目识别
    type_hint = _read_type_hint_from_row5(ws, header_row, data_start_row)

    # 解析表头
    headers = []
    required_columns = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        text = str(cell.value or '').strip()
        if not text:
            continue
        is_required = text.endswith('*') or text.endswith('＊')
        clean_text = text.rstrip('*＊').strip()
        col_letter = get_column_letter(col_idx)
        headers.append({
            'index': col_idx - 1,
            'letter': col_letter,
            'header': clean_text,
            'is_required': is_required,
        })
        if is_required:
            required_columns.append({
                'index': col_idx - 1,
                'letter': col_letter,
                'header': clean_text,
            })

    # 解析数据验证规则（下拉列表等）
    validations = []
    if ws.data_validations and ws.data_validations.dataValidation:
        for dv in ws.data_validations.dataValidation:
            validations.append({
                'ranges': str(dv.sqref) if dv.sqref else '',
                'type': dv.type or '',
                'formula1': dv.formula1 or '',
                'allow_blank': dv.allow_blank,
            })

    # 计算结构哈希
    sorted_header_names = sorted([h['header'] for h in headers])
    schema_hash = hashlib.sha256(
        ','.join(sorted_header_names).encode('utf-8')
    ).hexdigest()[:16]

    wb.close()

    return {
        'schema_hash': schema_hash,
        'headers': headers,
        'required_columns': required_columns,
        'sheet_names': sheet_names,
        'data_start_row': data_start_row,
        'header_row': header_row,
        'validations': validations,
        'type_hint': type_hint,  # 第 5 行「类型*」列的示例值
    }


def _detect_header_row(ws):
    """检测表头行号（查找内容最多的行，默认回退到第 2 行）"""
    best_row = 2
    best_count = 0
    for r in range(1, min(4, ws.max_row + 1)):
        count = sum(1 for c in range(1, ws.max_column + 1) if ws.cell(row=r, column=c).value)
        if count > best_count:
            best_count = count
            best_row = r
    return best_row


def _detect_data_start_row(ws, header_row):
    """检测数据起始行（表头行之后第一个 A 列有示例数据或空的行）"""
    for r in range(header_row + 1, min(header_row + 10, ws.max_row + 1)):
        # 如果该行任何列有值，可能就是数据起始行
        has_content = any(
            ws.cell(row=r, column=c).value is not None
            for c in range(1, min(ws.max_column + 1, 50))
        )
        if has_content:
            return r
    return header_row + 3  # 默认跳到第 5 行 (header_row=2 + 3)


def _read_type_hint_from_row5(ws, header_row, data_start_row):
    """
    从模板第 5 行「类型*」列读取示例类型值（如 "取景器"、"Видоискатель"），
    用于辅助类目识别。
    """
    # 先找到「类型」列的索引
    type_col = None
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        text = str(cell.value or '').rstrip('*＊').strip()
        if text in ('类型', 'Тип'):
            type_col = col_idx
            break

    if type_col:
        val = ws.cell(row=data_start_row, column=type_col).value
        if val:
            return str(val).strip()
    return None


# ═══════════════════════════════════════════════════════════════
# HTML 处理 & 富内容校验
# ═══════════════════════════════════════════════════════════════

import re as _re


def html_to_plain_text(html, max_length=3000):
    """
    从 OZON PDP HTML 提取纯文本俄语简介。
    移除 script/style/img 标签，提取可读文本，压缩空白，截断到 max_length。
    """
    if not html:
        return ''

    try:
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(html, 'html.parser')
        # 移除不需要的标签
        for tag in soup.find_all(['script', 'style', 'img', 'svg', 'noscript']):
            tag.decompose()
        text = soup.get_text(separator='\n')
    except Exception:
        # BeautifulSoup 不可用时，用正则简单处理
        text = _re.sub(r'<script[^>]*>.*?</script>', '', html, flags=_re.DOTALL | _re.IGNORECASE)
        text = _re.sub(r'<style[^>]*>.*?</style>', '', text, flags=_re.DOTALL | _re.IGNORECASE)
        text = _re.sub(r'<[^>]+>', ' ', text)
        text = _re.sub(r'&[a-z]+;', ' ', text)

    # 移除 "Описание" 等常见标题前缀
    text = _re.sub(r'\bОписание\b', '', text, flags=_re.IGNORECASE)
    # 压缩空白
    text = _re.sub(r'[ \t]+', ' ', text)
    text = _re.sub(r'\n{3,}', '\n\n', text)
    text = text.strip()

    # 截断
    if len(text) > max_length:
        # 尝试在句子边界截断
        cutoff = text.rfind('.', max_length - 200, max_length)
        if cutoff > max_length // 2:
            text = text[:cutoff + 1]
        else:
            text = text[:max_length]

    return text.strip()


def is_empty_rich_content(raw):
    """判断富内容 JSON 是否为空/无效"""
    if not raw:
        return True
    try:
        import json
        data = json.loads(raw) if isinstance(raw, str) else raw
    except (json.JSONDecodeError, TypeError):
        return True
    if not isinstance(data, dict):
        return True
    blocks = data.get('blocks', [])
    return len(blocks) == 0


def build_rich_content_from_draft(draft):
    """
    从草稿现有数据兜底生成 OZON 富内容 JSON。
    优先级：非空 rich_content_json > description_ru HTML 转 blocks

    返回:
        JSON 字符串或 ''
    """
    # 从 description_ru HTML 提取文本段落
    raw_desc = draft.description_ru or ''
    if not raw_desc:
        return ''

    try:
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(raw_desc, 'html.parser')
        for tag in soup.find_all(['script', 'style', 'noscript']):
            tag.decompose()

        blocks = []
        block_id = 0

        # 提取段落文本为 text block
        text_parts = []
        for p in soup.find_all(['p', 'div', 'li', 'h2', 'h3']):
            text = p.get_text(strip=True)
            if text and len(text) > 15:
                # 跳过纯数字/符号/英文标签
                if _re.match(r'^[0-9\s\.\-\+\*\#,;:!?\(\)\[\]{}<>/\\@&%"\'`~|]+$', text):
                    continue
                text = _re.sub(r'^\s*(Описание|Характеристики|Комплектация)\s*', '', text, flags=_re.IGNORECASE).strip()
                if text:
                    text_parts.append(text)

        if text_parts:
            # 合并太短的段落
            merged = []
            buf = ''
            for t in text_parts:
                if len(buf) + len(t) < 500:
                    buf += ('\n' if buf else '') + t
                else:
                    if buf:
                        merged.append(buf)
                    buf = t
            if buf:
                merged.append(buf)

            MAX_TEXT_BLOCKS = 5
            for paragraph in merged[:MAX_TEXT_BLOCKS]:  # 最多 5 个文本块
                block_id += 1
                if len(paragraph) > 2000:
                    paragraph = paragraph[:1997] + '...'
                blocks.append({
                    'id': block_id,
                    'type': 'text',
                    'content': {
                        'text': paragraph,
                        'size': 'body',
                        'align': 'left',
                    }
                })

        # 提取图片 URL
        from blueprints.ozon import _load_media_json
        media = _load_media_json(draft)
        media_images = media.get('images', []) if isinstance(media, dict) else []
        enabled_images = [i for i in media_images if i.get('selected') and i.get('role') == 'main']
        enabled_images.sort(key=lambda x: x.get('sort_order', 0))

        # 也从 HTML 提取图片
        html_img_urls = []
        for img_tag in soup.find_all('img'):
            src = img_tag.get('src') or img_tag.get('data-src') or ''
            if src and (src.startswith('http://') or src.startswith('https://')):
                html_img_urls.append(src)

        # 合并去重
        MAX_IMAGE_BLOCKS = 10
        seen_urls = set()
        for img in enabled_images[:MAX_IMAGE_BLOCKS]:
            url = img.get('ozon_url') or img.get('public_url') or img.get('url') or ''
            if url and url not in seen_urls:
                seen_urls.add(url)
                block_id += 1
                blocks.append({
                    'id': block_id,
                    'type': 'image',
                    'source': url,
                })

        for url in html_img_urls:
            if url not in seen_urls and len(seen_urls) < MAX_IMAGE_BLOCKS:
                seen_urls.add(url)
                block_id += 1
                blocks.append({
                    'id': block_id,
                    'type': 'image',
                    'source': url,
                })

        if blocks:
            import json
            return json.dumps({'version': '1.0', 'blocks': blocks}, ensure_ascii=False)
    except Exception:
        pass

    return ''


# ═══════════════════════════════════════════════════════════════
# 字段映射 — 草稿数据 → Excel 表头
# ═══════════════════════════════════════════════════════════════

def build_field_mapping(draft):
    """
    从草稿提取数据，映射到 OZON 模板列名。

    返回 dict: { "货号": "DJI-20260701-XXXX", "商品名称": "...", ... }
    """
    from models import OzonDraftSku
    from blueprints.ozon import _load_media_json, _load_draft_attributes_map, _safe_json_loads

    mapping = {}

    # ── 货号（第一行 SKU 的 offer_id）──
    first_sku = draft.draft_skus.order_by(OzonDraftSku.source_order).first()
    mapping['货号'] = (first_sku.offer_id or '').strip() if first_sku else ''

    # ── 商品名称 ──
    mapping['商品名称'] = (draft.title_ru or '').strip()

    # ── 价格 ──
    pricing = _safe_json_loads(draft.pricing_json, {})
    listing_price = pricing.get('listing_price', '') if isinstance(pricing, dict) else ''
    mapping['价格'] = listing_price
    mapping['价格,CNY'] = listing_price
    mapping['价格，CNY'] = listing_price

    # ── 主图链接 + 附加图片链接 ──
    image_urls = get_public_image_urls(draft)
    mapping['主图链接'] = image_urls[0] if image_urls else ''
    mapping['附加图片链接'] = '\n'.join(image_urls[1:]) if len(image_urls) > 1 else ''

    # ── 品牌 ──
    mapping['品牌'] = _extract_attribute_value(draft, 'brand', '品牌', 'Бренд')

    # ── 型号名称（优先从属性取 4180/9048，其次 source model）──
    model = _extract_attribute_value(draft, 'model', '型号', 'Модель')
    if not model and first_sku:
        model = (first_sku.source_sku_name or '').strip()
    if not model and draft.source:
        try:
            raw = json.loads(draft.source.raw_json or '{}')
        except (json.JSONDecodeError, TypeError):
            raw = {}
        model = (raw.get('model') or '').strip()
    mapping['型号名称'] = model

    # ── 类型（优先俄语值）──
    mapping['类型'] = (draft.type_name_ru or draft.type_name_cn or '').strip()

    # ── 简介（纯文本俄语描述，禁止 HTML）──
    description_text = ''
    raw_desc = draft.description_ru or ''
    # 判断是否为 HTML（含标签）
    if raw_desc and ('<' in raw_desc and '>' in raw_desc):
        description_text = html_to_plain_text(raw_desc)
    else:
        description_text = raw_desc.strip()
    # 如果 HTML 提取结果太短，用 bullets_ru 补充
    if len(description_text) < 50 and draft.bullets_ru:
        try:
            bullets = json.loads(draft.bullets_ru)
            if isinstance(bullets, list):
                description_text = '\n'.join(bullets)
        except (json.JSONDecodeError, TypeError):
            pass
    mapping['简介'] = description_text

    # ── JSON 富内容（优先 draft.rich_content_json，空则从 HTML+图片兜底生成）──
    rich = draft.rich_content_json or ''
    if is_empty_rich_content(rich):
        rich = build_rich_content_from_draft(draft)
    mapping['JSON富内容'] = rich

    # ── 商品颜色（属性优先，其次 SKU 颜色）──
    color = _extract_attribute_value(draft, 'color', '颜色', 'Цвет')
    if not color and first_sku:
        color = (first_sku.color_ru or '').strip()
    mapping['商品颜色'] = color

    # ── 毛重、包装尺寸（从属性中提取）──
    weight = _extract_attribute_value(draft, 'weight', '毛重', '重量', 'Вес', 'вес')
    mapping['毛重'] = weight
    mapping['毛重,克'] = weight
    mapping['毛重/克'] = weight
    mapping['毛重，克'] = weight
    mapping['包装宽度'] = _extract_attribute_value(draft, 'width', '宽度', 'Ширина', 'ширина')
    mapping['包装高度'] = _extract_attribute_value(draft, 'height', '高度', 'Высота', 'высота')
    mapping['包装长度'] = _extract_attribute_value(draft, 'length', '长度', 'Длина', 'длина')
    # 尝试从尺寸字符串解析宽/高/长
    size_str = _extract_attribute_value(draft, 'size', '尺寸', 'Размер')
    if size_str and (not mapping['包装宽度'] or not mapping['包装高度']):
        parts = size_str.replace('*', 'x').replace('X', 'x').split('x')
        if len(parts) == 3:
            if not mapping['包装宽度']:
                mapping['包装宽度'] = parts[0].strip()
            if not mapping['包装高度']:
                mapping['包装高度'] = parts[1].strip()
            if not mapping['包装长度']:
                mapping['包装长度'] = parts[2].strip()
    # 原产国：优先俄语
    country = _extract_attribute_value(draft, 'country', '原产国', '制造国', 'Страна', 'Китай')
    mapping['原产国'] = country if country else 'Китай'

    return mapping


def _extract_attribute_value(draft, *keywords):
    """
    从草稿属性中按关键词匹配提取值。

    匹配策略：
      1. 遍历 draft.attributes_json（已归一化）中的 attribute_id
      2. 用 attribute_id 查 OzonCategoryAttribute 获取 name/name_cn
      3. 关键词匹配属性名后，优先取俄语原值（字典值查 OzonAttributeValue.value）
    """
    from blueprints.ozon import _load_draft_attributes_map
    from models import OzonCategoryAttribute, OzonAttributeValue

    attrs_map = _load_draft_attributes_map(draft)
    if not attrs_map:
        return ''

    keywords_lower = [k.lower() for k in keywords]

    for attr_id_str, attr_data in attrs_map.items():
        if not isinstance(attr_data, dict):
            continue

        # 查 OzonCategoryAttribute 获取属性名
        cat_attr = (OzonCategoryAttribute
                    .select()
                    .where((OzonCategoryAttribute.user == draft.user) &
                           (OzonCategoryAttribute.attribute_id == str(attr_id_str)))
                    .first())

        attr_name = ''
        if cat_attr:
            attr_name = (cat_attr.name_cn or cat_attr.name or '').lower()
        inline_name = (attr_data.get('attribute_name') or '').lower()
        combined = (attr_name + ' ' + inline_name).strip()

        if not any(kw in combined for kw in keywords_lower):
            continue

        # 字典属性：查 OzonAttributeValue 取俄语原值
        if cat_attr and cat_attr.is_dictionary:
            value_id = attr_data.get('value_id') or attr_data.get('dict_value_id') or ''
            if value_id:
                dv = (OzonAttributeValue
                      .select()
                      .where((OzonAttributeValue.user == draft.user) &
                             (OzonAttributeValue.attribute_id == str(attr_id_str)) &
                             (OzonAttributeValue.value_id == str(value_id)))
                      .first())
                if dv and dv.value:
                    return dv.value.strip()

        # 非字典 / 字典值未找到：直接用 attributes_json 中的 value
        v = attr_data.get('value_ru') or attr_data.get('value') or ''
        if v:
            return str(v).strip()

    return ''


def get_public_image_url(draft):
    """
    从草稿媒体池获取可公开访问的主图 URL。
    优先 ozon_url（OZON CDN），其次 public_url（必须以 http/https 开头）。
    本地 /static/... 路径 OZON 无法读取，直接跳过。
    """
    from blueprints.ozon import _load_media_json
    media = _load_media_json(draft) if _load_media_json else {}
    images = media.get('images', []) if isinstance(media, dict) else []
    main_imgs = [i for i in images if i.get('selected') and i.get('role') == 'main']
    main_imgs.sort(key=lambda i: i.get('sort_order', 0))
    for img in main_imgs:
        url = img.get('ozon_url') or img.get('public_url') or img.get('url') or ''
        if url and (url.startswith('http://') or url.startswith('https://')):
            return url
    return None


def get_public_image_urls(draft):
    """
    从草稿媒体池获取所有已选主图的公开 URL 列表（第1张是封面，其余是附加图）。
    OZON 商品图集 = 主图链接 + 附加图片链接，不是富内容里的图片。
    """
    from blueprints.ozon import _load_media_json
    media = _load_media_json(draft) if _load_media_json else {}
    images = media.get('images', []) if isinstance(media, dict) else []
    main_imgs = [i for i in images if i.get('selected') and i.get('role') == 'main']
    main_imgs.sort(key=lambda i: i.get('sort_order', 0))
    urls = []
    for img in main_imgs:
        url = img.get('ozon_url') or img.get('public_url') or img.get('url') or ''
        if url and (url.startswith('http://') or url.startswith('https://')) and url not in urls:
            urls.append(url)
    return urls


# ═══════════════════════════════════════════════════════════════
# Excel 生成
# ═══════════════════════════════════════════════════════════════

def generate_export_excel(draft, template, field_mapping):
    """
    复制 OZON 官方模板，填入草稿数据，返回生成的 Excel 文件路径。

    参数:
        draft: OzonDraft 实例
        template: OzonExcelTemplate 实例
        field_mapping: build_field_mapping() 返回的字段→值映射 dict

    返回:
        (save_path, validation_errors) 元组
        - save_path: 生成的 .xlsx 文件绝对路径
        - validation_errors: 校验失败字段列表（为空表示通过）

    流程:
        1. 打开原始模板 workbook
        2. 清空数据区（第 5 行起），保留样式
        3. 按表头名匹配列 → 覆盖写入 field_mapping 值
        4. 保存后重新打开校验关键字段
    """
    from models import OzonDraftSku

    wb = openpyxl.load_workbook(template.stored_path)

    # 确保有 "模板" sheet
    if '模板' not in wb.sheetnames:
        wb.close()
        raise ValueError("模板文件缺少「模板」工作表")

    ws = wb['模板']
    data_start_row = template.data_start_row or 5
    header_row = template.header_row or 2

    # 构建列映射: header_text → column_index (1-based)
    col_map = {}
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        text = str(cell.value or '').rstrip('*＊').strip()
        if text:
            col_map[text] = col_idx

    # ── 清空数据区（第 5 行起），保留第 1-4 行说明不变 ──
    for r in range(data_start_row, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            # 保留公式（以 = 开头）
            if isinstance(cell.value, str) and cell.value.startswith('='):
                continue
            cell.value = None

    # 准备 SKU 列表（至少填充第一行）
    skus = list(draft.draft_skus.order_by(OzonDraftSku.source_order))
    if not skus:
        skus = [None]

    # ── 按行写入（覆盖模式）──
    row = data_start_row
    for sku in skus:
        row_data = dict(field_mapping)

        if sku:
            row_data['货号'] = (sku.offer_id or '').strip()
            row_data['SKU'] = (sku.offer_id or '').strip()
            # SKU 颜色：有值才覆盖（不覆盖属性提取的俄语值）
            sku_color = (getattr(sku, 'color_ru', None) or '').strip()
            if sku_color:
                row_data['商品颜色'] = sku_color
            if len(skus) > 1 and sku.offer_id:
                row_data['货号'] = (sku.offer_id or '').strip()

        for header_text, value in row_data.items():
            col_idx = col_map.get(header_text)
            if not col_idx:
                col_idx = _fuzzy_find_column(col_map, header_text)

            if col_idx and value:
                cell = ws.cell(row=row, column=col_idx)
                # 覆盖模式：第 5 行起强制写入（已清空，但以防万一）
                cell.value = value

        row += 1

    # ── 保存 ──
    save_dir = os.path.join(
        'uploads', 'ozon_template_exports',
        str(draft.user_id), str(draft.id)
    )
    os.makedirs(save_dir, exist_ok=True)
    ts = int(datetime.datetime.now().timestamp())
    first_sku = skus[0] if skus else None
    if first_sku and hasattr(first_sku, 'offer_id') and first_sku.offer_id:
        offer_id_slug = str(first_sku.offer_id).replace('/', '_')
    else:
        offer_id_slug = f'draft{draft.id}'
    filename = f'{ts}_{offer_id_slug}.xlsx'
    save_path = os.path.join(save_dir, filename)

    wb.save(save_path)
    wb.close()

    # ── 生成后校验 ──
    validation_errors = _validate_generated_excel(save_path, header_row, data_start_row)

    return save_path, validation_errors


def _validate_generated_excel(save_path, header_row, data_start_row):
    """
    重新打开生成的 Excel，校验关键字段是否填写。
    返回缺失字段名列表。
    """
    required_checks = {
        '货号': '货号*为空',
        '商品名称': '商品名称为空',
        '价格': '价格为空',
        '主图链接': '主图链接为空或非http链接',
        '品牌': '品牌*为空',
        '型号名称': '型号名称*为空',
        '类型': '类型*为空或包含中文',
    }

    errors = []
    try:
        wb = openpyxl.load_workbook(save_path, data_only=True)
        if '模板' not in wb.sheetnames:
            return ['模板 sheet 丢失']

        ws = wb['模板']
        col_map = {}
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            text = str(cell.value or '').rstrip('*＊').strip()
            if text:
                col_map[text] = col_idx

        row = data_start_row
        for check_header, error_msg in required_checks.items():
            col_idx = col_map.get(check_header)
            if not col_idx:
                col_idx = _fuzzy_find_column(col_map, check_header)
            if not col_idx:
                continue  # 模板没有此列，跳过

            val = ws.cell(row=row, column=col_idx).value
            val_str = str(val or '').strip()

            if check_header == '主图链接':
                if not val_str or not (val_str.startswith('http://') or val_str.startswith('https://')):
                    errors.append(error_msg)
            elif check_header == '类型':
                if not val_str:
                    errors.append(error_msg)
                else:
                    # 检查是否包含中文
                    has_cjk = any('一' <= ch <= '鿿' for ch in val_str)
                    if has_cjk:
                        errors.append('类型*包含中文: ' + val_str)
            else:
                if not val_str:
                    errors.append(error_msg)

        # 额外检查：价格列
        price_col = col_map.get('价格') or col_map.get('价格，CNY') or col_map.get('价格,CNY')
        if price_col:
            price_val = ws.cell(row=row, column=price_col).value
            if not price_val:
                errors.append('价格为空')

        # 简介列：不应包含 HTML 标签
        desc_col = col_map.get('简介')
        if desc_col:
            desc_val = str(ws.cell(row=row, column=desc_col).value or '')
            if '<div' in desc_val or '<span' in desc_val or '<img' in desc_val:
                errors.append('简介包含HTML标签，应为纯文本')

        # JSON富内容列：不应为空 blocks
        rich_col = col_map.get('JSON富内容')
        if rich_col:
            rich_val = str(ws.cell(row=row, column=rich_col).value or '')
            if not rich_val:
                errors.append('JSON富内容为空（请在草稿中先生成俄语富内容）')
            elif is_empty_rich_content(rich_val):
                errors.append('JSON富内容为空(blocks=[])，请先生成俄语富内容再导出')

        # 附加图片链接：有 >1 张主图时不应为空
        main_col = col_map.get('主图链接')
        extra_col = col_map.get('附加图片链接')
        if main_col and extra_col:
            main_val = str(ws.cell(row=row, column=main_col).value or '').strip()
            extra_val = str(ws.cell(row=row, column=extra_col).value or '').strip()
            if main_val and not extra_val:
                main_count = main_val.count('http')
                if main_count <= 1:
                    pass  # 只有 1 张主图，附加图为空是正常的
                else:
                    errors.append('有多个主图链接但附加图片链接为空，可能只上传了封面图')

        wb.close()
    except Exception as e:
        errors.append(f'校验失败: {str(e)}')

    return errors


def _fuzzy_find_column(col_map, header_text):
    """模糊匹配列：关键词出现在表头中，或表头关键词出现在目标中"""
    ht_lower = header_text.lower()
    for h, idx in col_map.items():
        if ht_lower in h.lower() or h.lower() in ht_lower:
            return idx
    # 特殊处理：价格列有多种写法
    if '价格' in ht_lower:
        for h, idx in col_map.items():
            if '价格' in h:
                return idx
    return None


# ═══════════════════════════════════════════════════════════════
# 类目自动识别 — 模板表头 → dcid/type_id
# ═══════════════════════════════════════════════════════════════

# 所有 OZON 模板共用的通用列名（不参与类目匹配）
_GENERIC_HEADERS = {
    '№', '货号', '商品名称', '商品名称（俄语）', '商品名称（英语）',
    '价格', '价格，cny', '价格，rub', '价格,cny', '价格,rub',
    '折扣前价格', '折扣前价格，cny', '折扣前价格，rub',
    '加速评价收集', 'sku', '条形码', '条形码（序列号/ean）',
    '毛重', '毛重，克', '毛重,克', '毛重/克',
    '包装宽度', '包装宽度，毫米', '包装高度', '包装高度，毫米', '包装长度', '包装长度，毫米',
    '主图链接', '附加图片链接', '照片货号',
    '品牌', '型号名称', '型号名称（针对合并为一张商品卡片）',
    '商品颜色', '类型', '原产国', '统一计量单位中的商品数量',
    '#主题标签', '简介', 'json富内容', '组合成类似的产品',
    '是否有序列号', '货号（其他变体）', '商品名称（英语）', '商品名称（俄语）',
    '价格，rub（包括增值税）', '增值税税率',
}

import re

def _normalize_attr_text(text):
    """归一化属性文本：小写、去标点、去空格"""
    text = text.lower().strip()
    # 去掉所有标点符号（中文和英文）
    text = re.sub(r'[，,。.；;：:！!？?（）()【】\[\]《》〈〉\-—/\s*＊#]+', '', text)
    return text


def _text_similarity(a, b):
    """
    计算两个归一化后文本的相似度 (0.0 ~ 1.0)。
    对中文文本使用字符级 Jaccard 相似度 + 互相包含检测。
    """
    if a == b:
        return 1.0
    if len(a) >= 2 and len(b) >= 2 and (a in b or b in a):
        return 0.85
    # 字符级 Jaccard
    sa, sb = set(a), set(b)
    if not sa or not sb:
        return 0.0
    intersection = sa & sb
    union = sa | sb
    jaccard = len(intersection) / len(union) if union else 0.0
    # 如果交集覆盖了较短文本的大部分字符
    shorter = min(len(sa), len(sb))
    coverage = len(intersection) / shorter if shorter else 0.0
    return max(jaccard, coverage * 0.8)


def identify_category_from_headers(headers, user_id):
    """
    用模板表头列名匹配 OzonCategoryAttribute 属性名，推断 dcid/type_id。

    参数:
        headers: inspect_template() 返回的 headers 列表
                 [{'header': '货号', 'is_required': True}, ...]
        user_id:  当前用户 ID

    返回:
        {'dcid': '...', 'type_id': '...', 'type_name': '...', 'confidence': 85}
        或 None（匹配置信度不足时）
    """
    from models import OzonCategoryAttribute, OzonCategoryType
    from collections import defaultdict
    from functools import reduce
    from peewee import operator as peewee_op

    # 清洗表头名集合，归一化 + 过滤通用列
    header_set = []
    for h in headers:
        text = h['header'].strip().rstrip('*＊')
        norm = _normalize_attr_text(text)
        if norm and text not in _GENERIC_HEADERS:
            header_set.append({'raw': text, 'norm': norm})

    if not header_set:
        return None

    # ══════════════════════════════════════════
    # 优化：先通过 SQL 预筛选候选 (dcid, type_id)
    # 只加载 name_cn 包含任意表头关键词的属性
    # ══════════════════════════════════════════
    candidate_keys = set()
    group = defaultdict(lambda: {'names': []})

    # 提取表头中有意义的关键词（取长度>=3的去重关键词）
    keywords = set()
    for h in header_set:
        norm = h['norm']
        if len(norm) >= 3:
            keywords.add(norm)

    candidate_keys = set()
    group = defaultdict(lambda: {'names': []})

    if keywords:
        # 阶段1：用少量关键词 LIKE 预筛选候选 (dcid, type_id)
        # 限制每批 LIKE 数量，分多批查询
        kw_list = sorted(keywords, key=lambda x: -len(x))[:20]  # 取最长的 20 个
        kw_chunks = [kw_list[i:i+5] for i in range(0, len(kw_list), 5)]  # 每批 5 个

        for chunk in kw_chunks:
            like_clauses = []
            for kw in chunk:
                like_clauses.append(OzonCategoryAttribute.name_cn.contains(kw))
            if like_clauses:
                cond = [OzonCategoryAttribute.user == user_id,
                        reduce(peewee_op.or_, like_clauses)]
                batch = (OzonCategoryAttribute
                         .select(OzonCategoryAttribute.ozon_category_id,
                                 OzonCategoryAttribute.type_id)
                         .where(*cond)
                         .distinct())
                candidate_keys.update((c.ozon_category_id, c.type_id or '') for c in batch)

    if not candidate_keys:
        return None

    # 限制候选数，取前 30 个类目组
    candidate_keys = set(list(candidate_keys)[:30])

    # 阶段2：分批加载候选类目的属性
    key_list = list(candidate_keys)
    for batch_start in range(0, len(key_list), 10):
        batch_keys = key_list[batch_start:batch_start+10]
        key_clauses = []
        for dcid, type_id in batch_keys:
            key_clauses.append(
                (OzonCategoryAttribute.ozon_category_id == dcid) &
                (OzonCategoryAttribute.type_id == type_id)
            )
        if key_clauses:
            cond2 = [OzonCategoryAttribute.user == user_id,
                     reduce(peewee_op.or_, key_clauses)]
            attrs = (OzonCategoryAttribute
                     .select(OzonCategoryAttribute.ozon_category_id,
                             OzonCategoryAttribute.type_id,
                             OzonCategoryAttribute.name,
                             OzonCategoryAttribute.name_cn)
                     .where(*cond2))
            for a in attrs:
                key = (a.ozon_category_id, a.type_id or '')
                raw_name = a.name_cn or a.name or ''
                norm = _normalize_attr_text(raw_name)
                if norm:
                    group[key]['names'].append({'raw': raw_name, 'norm': norm})
                if a.name:
                    norm_ru = _normalize_attr_text(a.name)
                    if norm_ru and norm_ru != norm:
                        group[key]['names'].append({'raw': a.name, 'norm': norm_ru})

    if not group:
        return None

    # 得分计算：对每个 (dcid, type_id)，计算所有表头的最佳匹配总分
    best_score = 0.0
    best_dcid = None
    best_type_id = None

    for (dcid, type_id), data in group.items():
        attr_names = data['names']
        if not attr_names:
            continue
        total = 0.0
        for h in header_set:
            best_sim = 0.0
            for a in attr_names:
                sim = _text_similarity(h['norm'], a['norm'])
                if sim > best_sim:
                    best_sim = sim
            if best_sim >= 0.5:
                total += best_sim
        score = total / len(header_set)
        if score > best_score:
            best_score = score
            best_dcid = dcid
            best_type_id = type_id

    if best_score < 0.35 or not best_dcid or not best_type_id:
        return None

    # 查找类型名称
    type_name = None
    if best_type_id:
        ct = (OzonCategoryType
              .select()
              .where((OzonCategoryType.user == user_id) &
                     (OzonCategoryType.description_category_id == best_dcid) &
                     (OzonCategoryType.type_id == best_type_id))
              .first())
        if ct:
            type_name = ct.type_name_cn or ct.type_name_ru

    return {
        'dcid': best_dcid,
        'type_id': best_type_id,
        'type_name': type_name,
        'confidence': min(round(best_score * 100), 99),
    }